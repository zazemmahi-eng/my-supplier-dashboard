# workspace_routes.py
"""
FastAPI routes for Workspace management.
Handles workspace CRUD, dataset upload with case-specific validation,
model selection, and KPI management.
"""

import io
import os
import uuid
import pandas as pd
from pathlib import Path
from datetime import datetime
from typing import Dict, Any, List, Optional
from fastapi import APIRouter, HTTPException, Depends, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy.orm.attributes import flag_modified
from pydantic import BaseModel, Field

from backend.database import get_db
from backend.workspace_models import (
    Workspace, WorkspaceDataset, CustomKPI, ModelSelection,
    DataTypeCase, WorkspaceStatus
)
from backend.mon_analyse import (
    calculer_kpis_globaux,
    calculer_risques_fournisseurs,
    obtenir_actions_recommandees,
    calculer_predictions_avancees,
    calculer_distribution_risques,
    comparer_methodes_prediction
)
# LLM-based ingestion module for intelligent column mapping
from backend.llm_ingestion import (
    analyze_csv_for_mapping,
    apply_mappings_and_normalize,
    process_csv_with_llm_mapping,
    ColumnRole
)

# Get backend directory for sample files
BACKEND_DIR = Path(__file__).resolve().parent

# ============================================
# ROUTER CONFIGURATION
# ============================================

router = APIRouter(prefix="/api/workspaces", tags=["workspaces"])

# ============================================
# PYDANTIC SCHEMAS
# ============================================

class WorkspaceCreate(BaseModel):
    """Schema for creating a new workspace"""
    name: str = Field(..., min_length=1, max_length=200)
    description: Optional[str] = None
    data_type: DataTypeCase = DataTypeCase.CASE_A


class WorkspaceUpdate(BaseModel):
    """Schema for updating workspace"""
    name: Optional[str] = Field(None, min_length=1, max_length=200)
    description: Optional[str] = None
    status: Optional[WorkspaceStatus] = None


class WorkspaceResponse(BaseModel):
    """Response schema for workspace"""
    id: uuid.UUID
    name: str
    description: Optional[str]
    data_type: str
    status: str
    created_at: datetime
    updated_at: datetime
    has_data: bool = False
    supplier_count: int = 0
    row_count: int = 0

    class Config:
        from_attributes = True


class CustomKPICreate(BaseModel):
    """Schema for creating custom KPI"""
    name: str = Field(..., min_length=1, max_length=100)
    description: Optional[str] = None
    formula_type: str = "average"  # average, sum, percentage, custom
    target_field: str = Field(..., min_length=1)
    threshold_warning: Optional[float] = None
    threshold_critical: Optional[float] = None
    unit: str = "%"
    decimal_places: int = 2


class ModelSelectionUpdate(BaseModel):
    """Schema for updating model selection"""
    selected_model: str = Field(..., pattern="^(moving_average|linear_regression|exponential|combined)$")
    parameters: Optional[Dict[str, Any]] = None


# ============================================
# DATA VALIDATION SCHEMAS PER CASE
# Each case has specific columns and generates case-specific dashboards
# ============================================

# Case A: DELAY ONLY
# For datasets containing ONLY delay data (no defects)
# Dashboard shows: delay KPIs, delay-based alerts, delay predictions
CASE_A_SCHEMA = {
    "required": ["supplier", "date_promised", "date_delivered"],
    "types": {
        "supplier": "string",
        "date_promised": "date",
        "date_delivered": "date"
    },
    "case_type": "delay_only",
    "description": "Données de retard uniquement (dates promises et livrées)"
}

# Case B: DEFECTS ONLY
# For datasets containing ONLY defect data (no delay)
# Dashboard shows: defects KPIs, defect-based alerts, defect predictions
CASE_B_SCHEMA = {
    "required": ["supplier", "order_date", "defects"],
    "types": {
        "supplier": "string",
        "order_date": "date",
        "defects": "float"
    },
    "case_type": "defects_only",
    "description": "Données de défauts uniquement (taux de défauts par commande)"
}

# Case C: MIXED (Delay + Defects)
# For datasets containing BOTH delay AND defects data
# Dashboard shows: all KPIs, combined alerts, predictions for both metrics
CASE_C_SCHEMA = {
    "required": ["supplier", "date_promised", "date_delivered", "defects"],
    "types": {
        "supplier": "string",
        "date_promised": "date",
        "date_delivered": "date",
        "defects": "float"
    },
    "case_type": "mixed",
    "description": "Données mixtes (retards ET défauts)"
}


# ============================================
# HELPER FUNCTIONS
# ============================================

def get_schema_for_case(data_type: DataTypeCase) -> Dict:
    """Returns the validation schema for a given data type case"""
    schemas = {
        DataTypeCase.CASE_A: CASE_A_SCHEMA,
        DataTypeCase.CASE_B: CASE_B_SCHEMA,
        DataTypeCase.CASE_C: CASE_C_SCHEMA
    }
    return schemas.get(data_type, CASE_A_SCHEMA)


def validate_csv_for_case(df: pd.DataFrame, data_type: DataTypeCase) -> List[str]:
    """
    Validates a DataFrame against the schema for a specific data type case.
    Returns list of error messages (empty if valid).
    """
    errors = []
    schema = get_schema_for_case(data_type)
    
    # Check required columns
    missing = [col for col in schema["required"] if col not in df.columns]
    if missing:
        errors.append(f"Colonnes manquantes pour Case {data_type.value}: {', '.join(missing)}")
        return errors
    
    # Validate data types
    for col, expected_type in schema["types"].items():
        if col not in df.columns:
            continue
            
        if expected_type == "date":
            try:
                pd.to_datetime(df[col], errors='raise')
            except Exception:
                errors.append(f"Colonne '{col}' contient des dates invalides. Format attendu: YYYY-MM-DD")
        
        elif expected_type == "float":
            try:
                pd.to_numeric(df[col], errors='raise')
            except Exception:
                errors.append(f"Colonne '{col}' doit contenir des nombres décimaux")
        
        elif expected_type == "integer":
            try:
                pd.to_numeric(df[col], errors='raise')
            except Exception:
                errors.append(f"Colonne '{col}' doit contenir des nombres entiers")
    
    # Check for empty supplier names
    if "supplier" in df.columns:
        if df["supplier"].isna().any() or (df["supplier"].astype(str).str.strip() == "").any():
            errors.append("Colonne 'supplier' contient des valeurs vides")
    
    return errors


def process_csv_for_case(df: pd.DataFrame, data_type: DataTypeCase) -> pd.DataFrame:
    """
    Process and normalize CSV data based on the data type case.
    Transforms data to be compatible with existing ML models WITHOUT modifying them.
    
    Case A (Delay Only): Calculates delay from dates, sets defects to 0
    Case B (Defects Only): Uses defects data, sets delay to 0
    Case C (Mixed): Uses both delay and defects data
    """
    processed_df = df.copy()
    
    if data_type == DataTypeCase.CASE_A:
        # ========================================
        # CASE A: DELAY ONLY
        # Accepts: supplier, date_promised, date_delivered
        # Dashboard: delay KPIs, delay alerts, delay predictions
        # ========================================
        processed_df["date_promised"] = pd.to_datetime(processed_df["date_promised"]).dt.tz_localize(None)
        processed_df["date_delivered"] = pd.to_datetime(processed_df["date_delivered"]).dt.tz_localize(None)
        # Calculate delay from dates
        processed_df["delay"] = (processed_df["date_delivered"] - processed_df["date_promised"]).dt.days
        processed_df["delay"] = processed_df["delay"].apply(lambda x: max(x, 0) if pd.notna(x) else 0)
        # Set defects to 0 for ML model compatibility (not used in Case A dashboard)
        processed_df["defects"] = 0.0
    
    elif data_type == DataTypeCase.CASE_B:
        # ========================================
        # CASE B: DEFECTS ONLY
        # Accepts: supplier, order_date, defects
        # Dashboard: defects KPIs, defect alerts, defect predictions
        # ========================================
        processed_df["order_date"] = pd.to_datetime(processed_df["order_date"]).dt.tz_localize(None)
        processed_df["defects"] = pd.to_numeric(processed_df["defects"], errors='coerce').fillna(0.0)
        # Set delay to 0 for ML model compatibility (not used in Case B dashboard)
        processed_df["delay"] = 0
        # Create date columns for compatibility with existing analysis functions
        processed_df["date_promised"] = processed_df["order_date"]
        processed_df["date_delivered"] = processed_df["order_date"]
    
    elif data_type == DataTypeCase.CASE_C:
        # ========================================
        # CASE C: MIXED (Delay + Defects)
        # Accepts: supplier, date_promised, date_delivered, defects
        # Dashboard: all KPIs, combined alerts, predictions for both
        # ========================================
        processed_df["date_promised"] = pd.to_datetime(processed_df["date_promised"]).dt.tz_localize(None)
        processed_df["date_delivered"] = pd.to_datetime(processed_df["date_delivered"]).dt.tz_localize(None)
        processed_df["defects"] = pd.to_numeric(processed_df["defects"], errors='coerce').fillna(0.0)
        # Calculate delay from dates
        processed_df["delay"] = (processed_df["date_delivered"] - processed_df["date_promised"]).dt.days
        processed_df["delay"] = processed_df["delay"].apply(lambda x: max(x, 0) if pd.notna(x) else 0)
    
    # Clean supplier names
    processed_df["supplier"] = processed_df["supplier"].astype(str).str.strip()
    
    # Sort by supplier and date
    date_col = "date_promised" if "date_promised" in processed_df.columns else "order_date"
    processed_df = processed_df.sort_values(["supplier", date_col]).reset_index(drop=True)
    
    return processed_df


def get_workspace_dataframe(workspace_id: uuid.UUID, db: Session) -> Optional[pd.DataFrame]:
    """
    Retrieves the active dataset for a workspace and returns it as a DataFrame.
    Ensures proper data types for ML model compatibility.
    Computes derived columns if missing (for backward compatibility with old data).
    """
    dataset = db.query(WorkspaceDataset).filter(
        WorkspaceDataset.workspace_id == workspace_id,
        WorkspaceDataset.is_active == True
    ).first()
    
    if not dataset or not dataset.data_json:
        return None
    
    df = pd.DataFrame(dataset.data_json)
    
    # Ensure proper data types for ML model compatibility
    # Convert date columns back to datetime
    for col in ['date_promised', 'date_delivered', 'order_date']:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors='coerce')
    
    # Compute delay if missing (backward compatibility)
    if 'delay' not in df.columns:
        if 'date_promised' in df.columns and 'date_delivered' in df.columns:
            df['delay'] = (df['date_delivered'] - df['date_promised']).dt.days
            df['delay'] = df['delay'].apply(lambda x: max(x, 0) if pd.notna(x) else 0)
        elif 'expected_days' in df.columns and 'actual_days' in df.columns:
            df['delay'] = df['actual_days'] - df['expected_days']
            df['delay'] = df['delay'].apply(lambda x: max(x, 0) if pd.notna(x) else 0)
        else:
            df['delay'] = 0  # Default to 0 if we can't compute
    
    # Ensure numeric columns are properly typed
    if 'delay' in df.columns:
        df['delay'] = pd.to_numeric(df['delay'], errors='coerce').fillna(0).astype(int)
    
    if 'defects' in df.columns:
        df['defects'] = pd.to_numeric(df['defects'], errors='coerce').fillna(0.0)
    
    if 'quality_score' in df.columns:
        df['quality_score'] = pd.to_numeric(df['quality_score'], errors='coerce').fillna(100.0)
    
    return df


# ============================================
# WORKSPACE CRUD ENDPOINTS
# ============================================

@router.get("", response_model=List[WorkspaceResponse])
async def list_workspaces(
    status: Optional[WorkspaceStatus] = None,
    user_id: Optional[str] = Query(None, description="Filter by user ID"),
    db: Session = Depends(get_db)
):
    """
    List all workspaces with their basic info.
    Optionally filter by status and/or user_id.
    """
    query = db.query(Workspace)
    
    if status:
        query = query.filter(Workspace.status == status)
    
    if user_id:
        query = query.filter(Workspace.owner_id == user_id)
    
    workspaces = query.order_by(Workspace.created_at.desc()).all()
    
    # Enrich with data status
    result = []
    for ws in workspaces:
        active_dataset = db.query(WorkspaceDataset).filter(
            WorkspaceDataset.workspace_id == ws.id,
            WorkspaceDataset.is_active == True
        ).first()
        
        result.append(WorkspaceResponse(
            id=ws.id,
            name=ws.name,
            description=ws.description,
            data_type=ws.data_type.value,
            status=ws.status.value,
            created_at=ws.created_at,
            updated_at=ws.updated_at,
            has_data=active_dataset is not None,
            supplier_count=len(active_dataset.suppliers) if active_dataset else 0,
            row_count=active_dataset.row_count if active_dataset else 0
        ))
    
    return result


# ============================================
# GLOBAL DASHBOARD AGGREGATION ENDPOINT
# ============================================

@router.get("/global/dashboard", response_model=Dict[str, Any])
async def get_global_dashboard(
    user_id: Optional[str] = Query(None, description="Filter by user ID"),
    db: Session = Depends(get_db)
):
    """
    Get aggregated dashboard data across ALL user workspaces.
    Returns high-level KPIs, workspace summaries, and global trends.
    This is a READ-ONLY overview - no data upload here.
    """
    # Get all workspaces (optionally filtered by user)
    query = db.query(Workspace).filter(Workspace.status == WorkspaceStatus.ACTIVE)
    if user_id:
        query = query.filter(Workspace.owner_id == user_id)
    
    workspaces = query.order_by(Workspace.created_at.desc()).all()
    
    # Initialize aggregated metrics
    total_workspaces = len(workspaces)
    total_suppliers = 0
    total_orders = 0
    workspaces_with_data = 0
    
    # Aggregated KPIs
    total_delay_sum = 0
    total_defect_sum = 0
    delay_count = 0
    defect_count = 0
    
    # Workspace summaries
    workspace_summaries = []
    
    # Risk distribution across all workspaces
    risk_distribution = {"faible": 0, "modere": 0, "eleve": 0}
    
    # Process each workspace
    for ws in workspaces:
        active_dataset = db.query(WorkspaceDataset).filter(
            WorkspaceDataset.workspace_id == ws.id,
            WorkspaceDataset.is_active == True
        ).first()
        
        supplier_count = len(active_dataset.suppliers) if active_dataset else 0
        row_count = active_dataset.row_count if active_dataset else 0
        has_data = active_dataset is not None
        
        if has_data:
            workspaces_with_data += 1
            total_suppliers += supplier_count
            total_orders += row_count
            
            # Get workspace dataframe for aggregation
            df = get_workspace_dataframe(ws.id, db)
            if df is not None and not df.empty:
                # Aggregate delays (if applicable)
                if 'delay' in df.columns and ws.data_type in [DataTypeCase.CASE_A, DataTypeCase.CASE_C]:
                    total_delay_sum += df['delay'].sum()
                    delay_count += len(df)
                
                # Aggregate defects (if applicable)
                if 'defects' in df.columns and ws.data_type in [DataTypeCase.CASE_B, DataTypeCase.CASE_C]:
                    total_defect_sum += df['defects'].sum()
                    defect_count += len(df)
                
                # Calculate risk distribution for this workspace
                try:
                    risques = calculer_risques_fournisseurs(df)
                    for r in risques:
                        niveau = r.get('niveau_risque', '').lower()
                        if 'faible' in niveau:
                            risk_distribution["faible"] += 1
                        elif 'modéré' in niveau or 'modere' in niveau:
                            risk_distribution["modere"] += 1
                        elif 'élevé' in niveau or 'eleve' in niveau:
                            risk_distribution["eleve"] += 1
                except:
                    pass
        
        # Case label mapping
        case_labels = {
            "delays": "Case A - Retards",
            "late_days": "Case B - Défauts",
            "mixed": "Case C - Mixte"
        }
        
        workspace_summaries.append({
            "id": str(ws.id),
            "name": ws.name,
            "description": ws.description,
            "data_type": ws.data_type.value,
            "case_label": case_labels.get(ws.data_type.value, ws.data_type.value),
            "status": ws.status.value,
            "has_data": has_data,
            "supplier_count": supplier_count,
            "row_count": row_count,
            "created_at": ws.created_at.isoformat(),
            "updated_at": ws.updated_at.isoformat()
        })
    
    # Calculate global averages
    avg_delay = (total_delay_sum / delay_count) if delay_count > 0 else 0
    avg_defect = (total_defect_sum / defect_count) if defect_count > 0 else 0
    
    return {
        "summary": {
            "total_workspaces": total_workspaces,
            "workspaces_with_data": workspaces_with_data,
            "total_suppliers": total_suppliers,
            "total_orders": total_orders,
            "unique_suppliers": total_suppliers  # Note: may have duplicates across workspaces
        },
        "global_kpis": {
            "avg_delay": round(avg_delay, 2),
            "avg_defect": round(avg_defect, 2),
            "delay_orders_analyzed": delay_count,
            "defect_orders_analyzed": defect_count
        },
        "risk_distribution": risk_distribution,
        "workspaces": workspace_summaries,
        "case_breakdown": {
            "case_a_count": len([ws for ws in workspace_summaries if ws["data_type"] == "delays"]),
            "case_b_count": len([ws for ws in workspace_summaries if ws["data_type"] == "late_days"]),
            "case_c_count": len([ws for ws in workspace_summaries if ws["data_type"] == "mixed"])
        }
    }


@router.post("", response_model=WorkspaceResponse)
async def create_workspace(
    workspace: WorkspaceCreate,
    db: Session = Depends(get_db)
):
    """
    Create a new workspace with specified name and data type.
    """
    # Check for duplicate name
    existing = db.query(Workspace).filter(Workspace.name == workspace.name).first()
    if existing:
        raise HTTPException(status_code=400, detail=f"Un workspace nommé '{workspace.name}' existe déjà")
    
    new_workspace = Workspace(
        name=workspace.name,
        description=workspace.description,
        data_type=workspace.data_type
    )
    
    db.add(new_workspace)
    db.commit()
    db.refresh(new_workspace)
    
    return WorkspaceResponse(
        id=new_workspace.id,
        name=new_workspace.name,
        description=new_workspace.description,
        data_type=new_workspace.data_type.value,
        status=new_workspace.status.value,
        created_at=new_workspace.created_at,
        updated_at=new_workspace.updated_at,
        has_data=False,
        supplier_count=0,
        row_count=0
    )


@router.get("/{workspace_id}", response_model=Dict[str, Any])
async def get_workspace(
    workspace_id: uuid.UUID,
    db: Session = Depends(get_db)
):
    """
    Get detailed workspace information including dataset and KPI status.
    """
    workspace = db.query(Workspace).filter(Workspace.id == workspace_id).first()
    
    if not workspace:
        raise HTTPException(status_code=404, detail="Workspace non trouvé")
    
    # Get active dataset
    active_dataset = db.query(WorkspaceDataset).filter(
        WorkspaceDataset.workspace_id == workspace_id,
        WorkspaceDataset.is_active == True
    ).first()
    
    # Get custom KPIs
    custom_kpis = db.query(CustomKPI).filter(
        CustomKPI.workspace_id == workspace_id,
        CustomKPI.is_enabled == True
    ).all()
    
    # Get model selection
    model_selection = db.query(ModelSelection).filter(
        ModelSelection.workspace_id == workspace_id
    ).first()
    
    return {
        "workspace": {
            "id": str(workspace.id),
            "name": workspace.name,
            "description": workspace.description,
            "data_type": workspace.data_type.value,
            "status": workspace.status.value,
            "created_at": workspace.created_at.isoformat(),
            "updated_at": workspace.updated_at.isoformat()
        },
        "dataset": {
            "has_data": active_dataset is not None,
            "filename": active_dataset.filename if active_dataset else None,
            "row_count": active_dataset.row_count if active_dataset else 0,
            "suppliers": active_dataset.suppliers if active_dataset else [],
            "date_range": {
                "start": active_dataset.date_start.isoformat() if active_dataset and active_dataset.date_start else None,
                "end": active_dataset.date_end.isoformat() if active_dataset and active_dataset.date_end else None
            } if active_dataset else None,
            "uploaded_at": active_dataset.uploaded_at.isoformat() if active_dataset else None
        },
        "custom_kpis": [
            {
                "id": str(kpi.id),
                "name": kpi.name,
                "formula_type": kpi.formula_type,
                "target_field": kpi.target_field,
                "unit": kpi.unit
            }
            for kpi in custom_kpis
        ],
        "model_selection": {
            "selected_model": model_selection.selected_model if model_selection else "combined",
            "parameters": model_selection.parameters if model_selection else {},
            "last_run_at": model_selection.last_run_at.isoformat() if model_selection and model_selection.last_run_at else None
        },
        "schema": get_schema_for_case(workspace.data_type)
    }


@router.put("/{workspace_id}", response_model=WorkspaceResponse)
async def update_workspace(
    workspace_id: uuid.UUID,
    update_data: WorkspaceUpdate,
    db: Session = Depends(get_db)
):
    """
    Update workspace details (name, description, status).
    """
    workspace = db.query(Workspace).filter(Workspace.id == workspace_id).first()
    
    if not workspace:
        raise HTTPException(status_code=404, detail="Workspace non trouvé")
    
    if update_data.name:
        # Check for duplicate name
        existing = db.query(Workspace).filter(
            Workspace.name == update_data.name,
            Workspace.id != workspace_id
        ).first()
        if existing:
            raise HTTPException(status_code=400, detail=f"Un workspace nommé '{update_data.name}' existe déjà")
        workspace.name = update_data.name
    
    if update_data.description is not None:
        workspace.description = update_data.description
    
    if update_data.status:
        workspace.status = update_data.status
    
    db.commit()
    db.refresh(workspace)
    
    active_dataset = db.query(WorkspaceDataset).filter(
        WorkspaceDataset.workspace_id == workspace_id,
        WorkspaceDataset.is_active == True
    ).first()
    
    return WorkspaceResponse(
        id=workspace.id,
        name=workspace.name,
        description=workspace.description,
        data_type=workspace.data_type.value,
        status=workspace.status.value,
        created_at=workspace.created_at,
        updated_at=workspace.updated_at,
        has_data=active_dataset is not None,
        supplier_count=len(active_dataset.suppliers) if active_dataset else 0,
        row_count=active_dataset.row_count if active_dataset else 0
    )


@router.delete("/{workspace_id}")
async def delete_workspace(
    workspace_id: uuid.UUID,
    db: Session = Depends(get_db)
):
    """
    Delete a workspace and all associated data.
    """
    workspace = db.query(Workspace).filter(Workspace.id == workspace_id).first()
    
    if not workspace:
        raise HTTPException(status_code=404, detail="Workspace non trouvé")
    
    db.delete(workspace)
    db.commit()
    
    return {"message": f"Workspace '{workspace.name}' supprimé avec succès"}


# ============================================
# DATASET UPLOAD ENDPOINTS
# ============================================

@router.post("/{workspace_id}/upload", response_model=Dict[str, Any])
async def upload_dataset(
    workspace_id: uuid.UUID,
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    """
    Upload and validate a CSV dataset for a specific workspace.
    Validates according to the workspace's data type case (A, B, or C).
    """
    # Get workspace
    workspace = db.query(Workspace).filter(Workspace.id == workspace_id).first()
    if not workspace:
        raise HTTPException(status_code=404, detail="Workspace non trouvé")
    
    # Check file type
    if not file.filename.endswith('.csv'):
        raise HTTPException(status_code=400, detail="Format invalide. Fichier CSV requis.")
    
    try:
        # Read CSV
        content = await file.read()
        df = pd.read_csv(io.BytesIO(content))
        
        if df.empty:
            raise HTTPException(status_code=400, detail="Le fichier CSV est vide.")
        
        # Validate against case-specific schema
        errors = validate_csv_for_case(df, workspace.data_type)
        if errors:
            raise HTTPException(
                status_code=400,
                detail={"message": "Erreurs de validation", "errors": errors}
            )
        
        # Process data according to case
        processed_df = process_csv_for_case(df, workspace.data_type)
        
        # Deactivate previous datasets
        db.query(WorkspaceDataset).filter(
            WorkspaceDataset.workspace_id == workspace_id
        ).update({"is_active": False})
        
        # Get date range
        date_col = "date_promised" if "date_promised" in processed_df.columns else "order_date"
        date_start = processed_df[date_col].min()
        date_end = processed_df[date_col].max()
        
        # Convert Pandas Timestamps to strings for JSON serialization
        df_for_json = processed_df.copy()
        for col in df_for_json.columns:
            if pd.api.types.is_datetime64_any_dtype(df_for_json[col]):
                df_for_json[col] = df_for_json[col].dt.strftime("%Y-%m-%d")
        
        # Convert date_start and date_end to Python datetime
        date_start_py = pd.to_datetime(date_start).to_pydatetime() if pd.notna(date_start) else None
        date_end_py = pd.to_datetime(date_end).to_pydatetime() if pd.notna(date_end) else None
        
        # Create new dataset record
        new_dataset = WorkspaceDataset(
            workspace_id=workspace_id,
            filename=file.filename,
            row_count=len(processed_df),
            column_count=len(processed_df.columns),
            suppliers=processed_df["supplier"].unique().tolist(),
            date_start=date_start_py,
            date_end=date_end_py,
            data_json=df_for_json.to_dict(orient='records'),
            is_active=True
        )
        
        db.add(new_dataset)
        db.commit()
        db.refresh(new_dataset)
        
        return {
            "success": True,
            "message": f"Dataset uploadé avec succès (Case {workspace.data_type.value})",
            "dataset_id": str(new_dataset.id),
            "summary": {
                "filename": file.filename,
                "total_rows": len(processed_df),
                "suppliers": len(new_dataset.suppliers),
                "supplier_list": new_dataset.suppliers,
                "date_range": {
                    "start": date_start_py.strftime("%Y-%m-%d") if date_start_py else None,
                    "end": date_end_py.strftime("%Y-%m-%d") if date_end_py else None
                }
            }
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur de traitement: {str(e)}")


# ============================================
# LLM-BASED CSV INGESTION ENDPOINTS
# Intelligent column mapping for arbitrary CSVs
# ============================================

class ColumnMappingRequest(BaseModel):
    """Schema for approving/editing column mappings"""
    source_column: str
    target_role: str  # supplier, date_promised, date_delivered, order_date, delay, defects, quality_score, ignore
    confidence: float = 1.0
    transformation_needed: Optional[str] = None


class ApplyMappingsRequest(BaseModel):
    """Schema for applying approved mappings"""
    mappings: List[ColumnMappingRequest]
    target_case: str  # delay_only, defects_only, mixed


@router.post("/{workspace_id}/upload/analyze", response_model=Dict[str, Any])
async def analyze_csv_for_llm_mapping(
    workspace_id: uuid.UUID,
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    """
    Step 1 of LLM-based ingestion: Analyze CSV and suggest column mappings.
    
    The LLM analyzes column names and sample data to suggest appropriate mappings.
    Returns mapping suggestions with confidence scores and detected issues.
    
    User can review and edit mappings before applying them.
    """
    # Validate workspace exists
    workspace = db.query(Workspace).filter(Workspace.id == workspace_id).first()
    if not workspace:
        raise HTTPException(status_code=404, detail="Workspace non trouvé")
    
    # Validate file type
    if not file.filename.endswith('.csv'):
        raise HTTPException(status_code=400, detail="Format invalide. Fichier CSV requis.")
    
    try:
        # Read CSV
        content = await file.read()
        df = pd.read_csv(io.BytesIO(content))
        
        if df.empty:
            raise HTTPException(status_code=400, detail="Le fichier CSV est vide.")
        
        # Analyze CSV with LLM-style column detection
        analysis = analyze_csv_for_mapping(df)
        
        # Store the raw CSV content temporarily in session/cache
        # For now, we'll return it encoded so frontend can send it back
        import base64
        csv_content_b64 = base64.b64encode(content).decode('utf-8')
        
        return {
            "success": True,
            "filename": file.filename,
            "row_count": len(df),
            "column_count": len(df.columns),
            "original_columns": list(df.columns),
            "analysis": analysis,
            "csv_content": csv_content_b64,  # Send back for step 2
            "workspace_data_type": workspace.data_type.value,
            "message": "Analyse terminée. Veuillez vérifier les mappings suggérés."
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur d'analyse: {str(e)}")


@router.post("/{workspace_id}/upload/apply-mappings", response_model=Dict[str, Any])
async def apply_llm_mappings(
    workspace_id: uuid.UUID,
    csv_content: str = Query(..., description="Base64 encoded CSV content"),
    mappings: str = Query(..., description="JSON string of approved mappings"),
    target_case: str = Query(..., description="Target case: delay_only, defects_only, mixed"),
    filename: str = Query("uploaded.csv", description="Original filename"),
    db: Session = Depends(get_db)
):
    """
    Step 2 of LLM-based ingestion: Apply approved mappings and normalize data.
    
    Takes the user-approved column mappings and applies all transformations:
    - Parses dates in multiple formats
    - Computes delay from dates if needed
    - Normalizes defects to 0-1 range
    - Validates all constraints
    
    All transformations are done by Python code, NOT the LLM.
    """
    import base64
    import json
    
    # Validate workspace exists
    workspace = db.query(Workspace).filter(Workspace.id == workspace_id).first()
    if not workspace:
        raise HTTPException(status_code=404, detail="Workspace non trouvé")
    
    try:
        # Decode CSV content
        csv_bytes = base64.b64decode(csv_content)
        df = pd.read_csv(io.BytesIO(csv_bytes))
        
        # Parse mappings
        approved_mappings = json.loads(mappings)
        
        # Map target_case to DataTypeCase
        case_mapping = {
            "delay_only": DataTypeCase.CASE_A,
            "defects_only": DataTypeCase.CASE_B,
            "mixed": DataTypeCase.CASE_C
        }
        data_type = case_mapping.get(target_case, workspace.data_type)
        
        # Apply mappings and normalize using the LLM ingestion module
        result = apply_mappings_and_normalize(df, approved_mappings, target_case)
        
        if not result.success:
            return {
                "success": False,
                "warnings": [w.message for w in result.warnings if w.severity == "warning"],
                "errors": [w.message for w in result.warnings if w.severity == "error"],
                "transformations": [t.details for t in result.transformations],
                "message": "La normalisation a échoué. Veuillez corriger les erreurs."
            }
        
        # Process with case-specific logic for backward compatibility
        processed_df = result.dataframe
        
        # Deactivate previous datasets
        db.query(WorkspaceDataset).filter(
            WorkspaceDataset.workspace_id == workspace_id
        ).update({"is_active": False})
        
        # Get date range
        date_col = "date_promised" if "date_promised" in processed_df.columns else "order_date"
        if date_col in processed_df.columns:
            date_start = processed_df[date_col].min()
            date_end = processed_df[date_col].max()
        else:
            date_start = date_end = None
        
        # Convert for JSON serialization
        df_for_json = processed_df.copy()
        for col in df_for_json.columns:
            if pd.api.types.is_datetime64_any_dtype(df_for_json[col]):
                df_for_json[col] = df_for_json[col].dt.strftime("%Y-%m-%d")
        
        # Convert dates
        date_start_py = pd.to_datetime(date_start).to_pydatetime() if pd.notna(date_start) else None
        date_end_py = pd.to_datetime(date_end).to_pydatetime() if pd.notna(date_end) else None
        
        # Update workspace data_type to match target case
        workspace.data_type = data_type
        
        # Create new dataset record
        new_dataset = WorkspaceDataset(
            workspace_id=workspace_id,
            filename=filename,
            row_count=len(processed_df),
            column_count=len(processed_df.columns),
            suppliers=processed_df["supplier"].unique().tolist() if "supplier" in processed_df.columns else [],
            date_start=date_start_py,
            date_end=date_end_py,
            data_json=df_for_json.to_dict(orient='records'),
            is_active=True
        )
        
        db.add(new_dataset)
        db.commit()
        db.refresh(new_dataset)
        
        return {
            "success": True,
            "message": f"Données normalisées et importées avec succès (Case: {target_case})",
            "dataset_id": str(new_dataset.id),
            "summary": result.summary,
            "transformations": [t.details for t in result.transformations],
            "warnings": [w.message for w in result.warnings if w.severity == "warning"],
            "detected_case": result.detected_case
        }
        
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Format de mappings invalide")
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Erreur de traitement: {str(e)}")


@router.post("/{workspace_id}/upload/smart", response_model=Dict[str, Any])
async def smart_upload_with_llm(
    workspace_id: uuid.UUID,
    file: UploadFile = File(...),
    auto_apply: bool = Query(False, description="Auto-apply high confidence mappings"),
    db: Session = Depends(get_db)
):
    """
    Combined LLM-based ingestion: Analyze and optionally auto-apply mappings.
    
    If auto_apply=True and all mappings have high confidence (>0.8), 
    automatically applies mappings and imports data.
    
    Otherwise, returns analysis for user review.
    """
    # Validate workspace exists
    workspace = db.query(Workspace).filter(Workspace.id == workspace_id).first()
    if not workspace:
        raise HTTPException(status_code=404, detail="Workspace non trouvé")
    
    # Validate file type
    if not file.filename.endswith('.csv'):
        raise HTTPException(status_code=400, detail="Format invalide. Fichier CSV requis.")
    
    try:
        # Read CSV
        content = await file.read()
        df = pd.read_csv(io.BytesIO(content))
        
        if df.empty:
            raise HTTPException(status_code=400, detail="Le fichier CSV est vide.")
        
        # Analyze CSV
        analysis = analyze_csv_for_mapping(df)
        
        # Check if all mappings have high confidence
        all_high_confidence = all(
            m["confidence"] > 0.8 
            for m in analysis["mappings"] 
            if m["target_role"] != "ignore"
        )
        
        no_errors = not any(i["severity"] == "error" for i in analysis.get("issues", []))
        
        if auto_apply and all_high_confidence and no_errors:
            # Auto-apply mappings
            result = process_csv_with_llm_mapping(
                df, 
                user_mappings=analysis["mappings"],
                target_case=analysis["detected_case"]
            )
            
            if result.success:
                # Save to database
                processed_df = result.dataframe
                
                # Deactivate previous datasets
                db.query(WorkspaceDataset).filter(
                    WorkspaceDataset.workspace_id == workspace_id
                ).update({"is_active": False})
                
                # Map detected case to DataTypeCase
                case_mapping = {
                    "delay_only": DataTypeCase.CASE_A,
                    "defects_only": DataTypeCase.CASE_B,
                    "mixed": DataTypeCase.CASE_C
                }
                data_type = case_mapping.get(result.detected_case, workspace.data_type)
                workspace.data_type = data_type
                
                # Get date range
                date_col = "date_promised" if "date_promised" in processed_df.columns else "order_date"
                if date_col in processed_df.columns:
                    date_start = processed_df[date_col].min()
                    date_end = processed_df[date_col].max()
                else:
                    date_start = date_end = None
                
                # Convert for JSON
                df_for_json = processed_df.copy()
                for col in df_for_json.columns:
                    if pd.api.types.is_datetime64_any_dtype(df_for_json[col]):
                        df_for_json[col] = df_for_json[col].dt.strftime("%Y-%m-%d")
                
                date_start_py = pd.to_datetime(date_start).to_pydatetime() if pd.notna(date_start) else None
                date_end_py = pd.to_datetime(date_end).to_pydatetime() if pd.notna(date_end) else None
                
                new_dataset = WorkspaceDataset(
                    workspace_id=workspace_id,
                    filename=file.filename,
                    row_count=len(processed_df),
                    column_count=len(processed_df.columns),
                    suppliers=processed_df["supplier"].unique().tolist() if "supplier" in processed_df.columns else [],
                    date_start=date_start_py,
                    date_end=date_end_py,
                    data_json=df_for_json.to_dict(orient='records'),
                    is_active=True
                )
                
                db.add(new_dataset)
                db.commit()
                db.refresh(new_dataset)
                
                return {
                    "success": True,
                    "auto_applied": True,
                    "message": f"Données importées automatiquement (Case: {result.detected_case})",
                    "dataset_id": str(new_dataset.id),
                    "summary": result.summary,
                    "transformations": [t.details for t in result.transformations],
                    "warnings": [w.message for w in result.warnings if w.severity == "warning"]
                }
        
        # Return analysis for manual review
        import base64
        csv_content_b64 = base64.b64encode(content).decode('utf-8')
        
        return {
            "success": True,
            "auto_applied": False,
            "needs_review": True,
            "filename": file.filename,
            "row_count": len(df),
            "column_count": len(df.columns),
            "original_columns": list(df.columns),
            "analysis": analysis,
            "csv_content": csv_content_b64,
            "message": "Certains mappings nécessitent une vérification manuelle."
        }
        
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Erreur de traitement: {str(e)}")


@router.get("/{workspace_id}/schema-info")
async def get_schema_info(
    workspace_id: uuid.UUID,
    db: Session = Depends(get_db)
):
    """
    Get the expected schema for the workspace's data type.
    Helps users understand what columns are required.
    """
    workspace = db.query(Workspace).filter(Workspace.id == workspace_id).first()
    if not workspace:
        raise HTTPException(status_code=404, detail="Workspace non trouvé")
    
    schema = get_schema_for_case(workspace.data_type)
    
    # Generate sample CSV content
    # Sample CSV templates for each data type case
    # Case A: delay only, Case B: defects only, Case C: mixed
    samples = {
        DataTypeCase.CASE_A: "supplier,date_promised,date_delivered\nSupplier A,2024-01-01,2024-01-03\nSupplier B,2024-01-05,2024-01-10",
        DataTypeCase.CASE_B: "supplier,order_date,defects\nSupplier A,2024-01-01,0.02\nSupplier B,2024-01-05,0.08",
        DataTypeCase.CASE_C: "supplier,date_promised,date_delivered,defects\nSupplier A,2024-01-01,2024-01-03,0.02\nSupplier B,2024-01-05,2024-01-10,0.08"
    }
    
    return {
        "data_type": workspace.data_type.value,
        "schema": schema,
        "sample_csv": samples.get(workspace.data_type, samples[DataTypeCase.CASE_A])
    }


@router.get("/{workspace_id}/sample-download")
async def download_sample_csv(
    workspace_id: uuid.UUID,
    db: Session = Depends(get_db)
):
    """
    Download a sample CSV file for the workspace's data type.
    """
    workspace = db.query(Workspace).filter(Workspace.id == workspace_id).first()
    if not workspace:
        raise HTTPException(status_code=404, detail="Workspace non trouvé")
    
    # Map data types to sample files
    sample_files = {
        DataTypeCase.CASE_A: "sample_data_case_a.csv",
        DataTypeCase.CASE_B: "sample_data_case_b.csv",
        DataTypeCase.CASE_C: "sample_data_case_c.csv"
    }
    
    sample_file = BACKEND_DIR / sample_files.get(workspace.data_type, "sample_data_case_a.csv")
    
    if not sample_file.exists():
        raise HTTPException(status_code=404, detail="Fichier exemple non trouvé")
    
    with open(sample_file, 'rb') as f:
        content = f.read()
    
    return StreamingResponse(
        io.BytesIO(content),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={sample_files[workspace.data_type]}"}
    )


# ============================================
# MODEL SELECTION ENDPOINTS
# ============================================

@router.get("/{workspace_id}/models")
async def list_available_models(workspace_id: uuid.UUID, db: Session = Depends(get_db)):
    """
    List all available ML models that can be used for predictions.
    These are the EXISTING models - no new models are created.
    """
    workspace = db.query(Workspace).filter(Workspace.id == workspace_id).first()
    if not workspace:
        raise HTTPException(status_code=404, detail="Workspace non trouvé")
    
    return {
        "models": [
            {
                "id": "moving_average",
                "name": "Moyenne Glissante",
                "description": "Utilise une fenêtre glissante pour calculer la moyenne des N dernières observations",
                "parameters": [
                    {"name": "fenetre", "type": "integer", "default": 3, "min": 1, "max": 10}
                ]
            },
            {
                "id": "linear_regression",
                "name": "Régression Linéaire",
                "description": "Modèle de régression pour détecter les tendances linéaires",
                "parameters": []
            },
            {
                "id": "exponential",
                "name": "Lissage Exponentiel",
                "description": "Pondère davantage les observations récentes avec un facteur alpha",
                "parameters": [
                    {"name": "alpha", "type": "float", "default": 0.3, "min": 0.1, "max": 0.9}
                ]
            },
            {
                "id": "combined",
                "name": "Modèle Combiné (Recommandé)",
                "description": "Combine les 3 méthodes pour une prédiction plus robuste",
                "parameters": [
                    {"name": "fenetre", "type": "integer", "default": 3, "min": 1, "max": 10}
                ]
            }
        ]
    }


@router.put("/{workspace_id}/model-selection")
async def update_model_selection(
    workspace_id: uuid.UUID,
    selection: ModelSelectionUpdate,
    db: Session = Depends(get_db)
):
    """
    Update the selected ML model for a workspace.
    Does NOT modify any model code - just stores the selection.
    """
    workspace = db.query(Workspace).filter(Workspace.id == workspace_id).first()
    if not workspace:
        raise HTTPException(status_code=404, detail="Workspace non trouvé")
    
    # Get or create model selection
    model_sel = db.query(ModelSelection).filter(
        ModelSelection.workspace_id == workspace_id
    ).first()
    
    if model_sel:
        model_sel.selected_model = selection.selected_model
        model_sel.parameters = selection.parameters or {}
        model_sel.cached_results = None  # Clear cache on model change
    else:
        model_sel = ModelSelection(
            workspace_id=workspace_id,
            selected_model=selection.selected_model,
            parameters=selection.parameters or {}
        )
        db.add(model_sel)
    
    db.commit()
    
    return {
        "success": True,
        "selected_model": selection.selected_model,
        "parameters": selection.parameters or {}
    }


# ============================================
# KPI MANAGEMENT ENDPOINTS
# ============================================

@router.get("/{workspace_id}/kpis")
async def get_kpis(workspace_id: uuid.UUID, db: Session = Depends(get_db)):
    """
    Get both standard and custom KPIs for a workspace.
    """
    workspace = db.query(Workspace).filter(Workspace.id == workspace_id).first()
    if not workspace:
        raise HTTPException(status_code=404, detail="Workspace non trouvé")
    
    # Standard KPIs (always available)
    standard_kpis = [
        {"id": "taux_retard", "name": "Taux de Retard", "unit": "%", "type": "standard"},
        {"id": "taux_defaut", "name": "Taux de Défaut", "unit": "%", "type": "standard"},
        {"id": "retard_moyen", "name": "Retard Moyen", "unit": "jours", "type": "standard"},
        {"id": "nb_commandes", "name": "Nombre de Commandes", "unit": "", "type": "standard"},
        {"id": "taux_conformite", "name": "Taux de Conformité", "unit": "%", "type": "standard"},
        {"id": "commandes_parfaites", "name": "Commandes Parfaites", "unit": "", "type": "standard"}
    ]
    
    # Custom KPIs from database
    custom_kpis = db.query(CustomKPI).filter(
        CustomKPI.workspace_id == workspace_id
    ).all()
    
    custom_kpi_list = [
        {
            "id": str(kpi.id),
            "name": kpi.name,
            "description": kpi.description,
            "formula_type": kpi.formula_type,
            "target_field": kpi.target_field,
            "unit": kpi.unit,
            "threshold_warning": kpi.threshold_warning,
            "threshold_critical": kpi.threshold_critical,
            "is_enabled": kpi.is_enabled,
            "type": "custom"
        }
        for kpi in custom_kpis
    ]
    
    return {
        "standard_kpis": standard_kpis,
        "custom_kpis": custom_kpi_list
    }


@router.post("/{workspace_id}/kpis/custom")
async def create_custom_kpi(
    workspace_id: uuid.UUID,
    kpi_data: CustomKPICreate,
    db: Session = Depends(get_db)
):
    """
    Create a custom KPI for a workspace.
    """
    workspace = db.query(Workspace).filter(Workspace.id == workspace_id).first()
    if not workspace:
        raise HTTPException(status_code=404, detail="Workspace non trouvé")
    
    new_kpi = CustomKPI(
        workspace_id=workspace_id,
        name=kpi_data.name,
        description=kpi_data.description,
        formula_type=kpi_data.formula_type,
        target_field=kpi_data.target_field,
        threshold_warning=kpi_data.threshold_warning,
        threshold_critical=kpi_data.threshold_critical,
        unit=kpi_data.unit,
        decimal_places=kpi_data.decimal_places
    )
    
    db.add(new_kpi)
    db.commit()
    db.refresh(new_kpi)
    
    return {
        "success": True,
        "kpi": {
            "id": str(new_kpi.id),
            "name": new_kpi.name,
            "formula_type": new_kpi.formula_type,
            "target_field": new_kpi.target_field,
            "unit": new_kpi.unit
        }
    }


@router.delete("/{workspace_id}/kpis/custom/{kpi_id}")
async def delete_custom_kpi(
    workspace_id: uuid.UUID,
    kpi_id: uuid.UUID,
    db: Session = Depends(get_db)
):
    """
    Delete a custom KPI.
    """
    kpi = db.query(CustomKPI).filter(
        CustomKPI.id == kpi_id,
        CustomKPI.workspace_id == workspace_id
    ).first()
    
    if not kpi:
        raise HTTPException(status_code=404, detail="KPI non trouvé")
    
    db.delete(kpi)
    db.commit()
    
    return {"success": True, "message": f"KPI '{kpi.name}' supprimé"}


# ============================================
# ANALYSIS ENDPOINTS (Using Existing Models)
# Case-specific dashboards: A=delay, B=defects, C=mixed
# ============================================

def calculate_case_specific_kpis(df: pd.DataFrame, data_type: DataTypeCase) -> Dict[str, Any]:
    """
    Calculate KPIs specific to the data type case.
    
    Case A (Delay Only): Only delay-related KPIs
    Case B (Defects Only): Only defects-related KPIs
    Case C (Mixed): All KPIs
    """
    kpis = {}
    
    if data_type == DataTypeCase.CASE_A:
        # ========================================
        # CASE A: DELAY ONLY KPIs
        # ========================================
        total_orders = len(df)
        delayed_orders = len(df[df['delay'] > 0])
        
        kpis['taux_retard'] = round((delayed_orders / total_orders * 100) if total_orders > 0 else 0, 2)
        kpis['retard_moyen'] = round(df['delay'].mean(), 1) if len(df) > 0 else 0
        kpis['retard_max'] = int(df['delay'].max()) if len(df) > 0 else 0
        kpis['nb_commandes'] = total_orders
        kpis['nb_retards'] = delayed_orders
        kpis['commandes_a_temps'] = total_orders - delayed_orders
        kpis['taux_ponctualite'] = round(100 - kpis['taux_retard'], 2)
        
    elif data_type == DataTypeCase.CASE_B:
        # ========================================
        # CASE B: DEFECTS ONLY KPIs
        # ========================================
        total_orders = len(df)
        defective_orders = len(df[df['defects'] > 0])
        
        kpis['taux_defaut'] = round((defective_orders / total_orders * 100) if total_orders > 0 else 0, 2)
        kpis['defaut_moyen'] = round(df['defects'].mean() * 100, 2) if len(df) > 0 else 0
        kpis['defaut_max'] = round(df['defects'].max() * 100, 2) if len(df) > 0 else 0
        kpis['nb_commandes'] = total_orders
        kpis['nb_defectueux'] = defective_orders
        kpis['commandes_conformes'] = total_orders - defective_orders
        kpis['taux_conformite'] = round(100 - kpis['taux_defaut'], 2)
        
    elif data_type == DataTypeCase.CASE_C:
        # ========================================
        # CASE C: MIXED KPIs (Both delay and defects)
        # ========================================
        total_orders = len(df)
        delayed_orders = len(df[df['delay'] > 0])
        defective_orders = len(df[df['defects'] > 0])
        perfect_orders = len(df[(df['delay'] == 0) & (df['defects'] == 0)])
        
        # Delay KPIs
        kpis['taux_retard'] = round((delayed_orders / total_orders * 100) if total_orders > 0 else 0, 2)
        kpis['retard_moyen'] = round(df['delay'].mean(), 1) if len(df) > 0 else 0
        
        # Defects KPIs
        kpis['taux_defaut'] = round((defective_orders / total_orders * 100) if total_orders > 0 else 0, 2)
        kpis['defaut_moyen'] = round(df['defects'].mean() * 100, 2) if len(df) > 0 else 0
        
        # Combined KPIs
        kpis['nb_commandes'] = total_orders
        kpis['commandes_parfaites'] = perfect_orders
        kpis['taux_conformite'] = round((perfect_orders / total_orders * 100) if total_orders > 0 else 0, 2)
    
    return kpis


def calculate_case_specific_supplier_risks(df: pd.DataFrame, data_type: DataTypeCase) -> List[Dict[str, Any]]:
    """
    Calculate supplier risks specific to the data type case.
    
    Case A: Risk based on delay only
    Case B: Risk based on defects only
    Case C: Risk based on both metrics
    """
    risques = []
    
    for supplier in df['supplier'].unique():
        supplier_df = df[df['supplier'] == supplier]
        n_orders = len(supplier_df)
        
        supplier_data = {
            'supplier': supplier,
            'nb_commandes': n_orders,
        }
        
        if data_type == DataTypeCase.CASE_A:
            # ========================================
            # CASE A: DELAY-BASED RISK
            # ========================================
            retard_moyen = supplier_df['delay'].mean()
            taux_retard = (supplier_df['delay'] > 0).mean() * 100
            
            # Risk score based on delay only (0-100)
            score_risque = min(100, int(retard_moyen * 5 + taux_retard * 0.5))
            
            supplier_data.update({
                'retard_moyen': round(retard_moyen, 1),
                'taux_retard': round(taux_retard, 1),
                'score_risque': score_risque,
                'niveau_risque': 'Élevé' if score_risque > 55 else 'Modéré' if score_risque > 25 else 'Faible',
                'status': 'eleve' if score_risque > 55 else 'modere' if score_risque > 25 else 'faible',
                'tendance_retards': 'hausse' if len(supplier_df) >= 3 and supplier_df['delay'].iloc[-1] > supplier_df['delay'].mean() else 'baisse'
            })
            
        elif data_type == DataTypeCase.CASE_B:
            # ========================================
            # CASE B: DEFECTS-BASED RISK
            # ========================================
            defaut_moyen = supplier_df['defects'].mean() * 100
            taux_defaut = (supplier_df['defects'] > 0).mean() * 100
            
            # Risk score based on defects only (0-100)
            score_risque = min(100, int(defaut_moyen * 2 + taux_defaut * 0.5))
            
            supplier_data.update({
                'taux_defaut': round(taux_defaut, 1),
                'defaut_moyen': round(defaut_moyen, 2),
                'score_risque': score_risque,
                'niveau_risque': 'Élevé' if score_risque > 55 else 'Modéré' if score_risque > 25 else 'Faible',
                'status': 'eleve' if score_risque > 55 else 'modere' if score_risque > 25 else 'faible',
                'tendance_defauts': 'hausse' if len(supplier_df) >= 3 and supplier_df['defects'].iloc[-1] > supplier_df['defects'].mean() else 'baisse'
            })
            
        elif data_type == DataTypeCase.CASE_C:
            # ========================================
            # CASE C: COMBINED RISK (delay + defects)
            # ========================================
            retard_moyen = supplier_df['delay'].mean()
            taux_retard = (supplier_df['delay'] > 0).mean() * 100
            defaut_moyen = supplier_df['defects'].mean() * 100
            taux_defaut = (supplier_df['defects'] > 0).mean() * 100
            
            # Combined risk score (0-100)
            score_risque = min(100, int(
                retard_moyen * 3 + 
                taux_retard * 0.3 + 
                defaut_moyen * 1.5 + 
                taux_defaut * 0.3
            ))
            
            supplier_data.update({
                'retard_moyen': round(retard_moyen, 1),
                'taux_retard': round(taux_retard, 1),
                'taux_defaut': round(taux_defaut, 1),
                'defaut_moyen': round(defaut_moyen, 2),
                'score_risque': score_risque,
                'niveau_risque': 'Élevé' if score_risque > 55 else 'Modéré' if score_risque > 25 else 'Faible',
                'status': 'eleve' if score_risque > 55 else 'modere' if score_risque > 25 else 'faible',
                'tendance_retards': 'hausse' if len(supplier_df) >= 3 and supplier_df['delay'].iloc[-1] > supplier_df['delay'].mean() else 'baisse',
                'tendance_defauts': 'hausse' if len(supplier_df) >= 3 and supplier_df['defects'].iloc[-1] > supplier_df['defects'].mean() else 'baisse'
            })
        
        risques.append(supplier_data)
    
    # Sort by risk score descending
    risques.sort(key=lambda x: x['score_risque'], reverse=True)
    return risques


def calculate_case_specific_actions(risques: List[Dict], data_type: DataTypeCase) -> List[Dict[str, Any]]:
    """
    Generate recommended actions specific to the data type case.
    
    Case A: Actions focused on delay reduction
    Case B: Actions focused on defect reduction
    Case C: Actions for both metrics
    """
    actions = []
    
    for r in risques:
        supplier = r['supplier']
        niveau = r['niveau_risque']
        
        if data_type == DataTypeCase.CASE_A:
            # ========================================
            # CASE A: DELAY-FOCUSED ACTIONS
            # ========================================
            if niveau == 'Élevé':
                actions.append({
                    'supplier': supplier,
                    'action': 'Renégocier les délais de livraison',
                    'priority': 'high',
                    'raison': f"Retard moyen de {r.get('retard_moyen', 0)} jours",
                    'delai': 'Immédiat',
                    'impact': 'Réduction des retards de 30-50%'
                })
            elif niveau == 'Modéré':
                actions.append({
                    'supplier': supplier,
                    'action': 'Mettre en place un suivi hebdomadaire des livraisons',
                    'priority': 'medium',
                    'raison': f"Taux de retard de {r.get('taux_retard', 0)}%",
                    'delai': '2 semaines',
                    'impact': 'Amélioration de la ponctualité'
                })
                
        elif data_type == DataTypeCase.CASE_B:
            # ========================================
            # CASE B: DEFECT-FOCUSED ACTIONS
            # ========================================
            if niveau == 'Élevé':
                actions.append({
                    'supplier': supplier,
                    'action': 'Audit qualité approfondi',
                    'priority': 'high',
                    'raison': f"Taux de défaut de {r.get('taux_defaut', 0)}%",
                    'delai': 'Immédiat',
                    'impact': 'Réduction des défauts de 40-60%'
                })
            elif niveau == 'Modéré':
                actions.append({
                    'supplier': supplier,
                    'action': 'Renforcer les contrôles qualité à réception',
                    'priority': 'medium',
                    'raison': f"Défaut moyen de {r.get('defaut_moyen', 0)}%",
                    'delai': '1 mois',
                    'impact': 'Amélioration de la conformité'
                })
                
        elif data_type == DataTypeCase.CASE_C:
            # ========================================
            # CASE C: COMBINED ACTIONS
            # ========================================
            if niveau == 'Élevé':
                # Check which metric is worse
                has_delay_issue = r.get('taux_retard', 0) > 30
                has_defect_issue = r.get('taux_defaut', 0) > 30
                
                if has_delay_issue and has_defect_issue:
                    actions.append({
                        'supplier': supplier,
                        'action': 'Plan d\'amélioration complet (délais + qualité)',
                        'priority': 'high',
                        'raison': f"Retard: {r.get('retard_moyen', 0)}j, Défaut: {r.get('taux_defaut', 0)}%",
                        'delai': 'Immédiat',
                        'impact': 'Amélioration globale de 40%'
                    })
                elif has_delay_issue:
                    actions.append({
                        'supplier': supplier,
                        'action': 'Renégocier les délais de livraison',
                        'priority': 'high',
                        'raison': f"Retard moyen de {r.get('retard_moyen', 0)} jours",
                        'delai': 'Immédiat',
                        'impact': 'Réduction des retards de 30-50%'
                    })
                else:
                    actions.append({
                        'supplier': supplier,
                        'action': 'Audit qualité approfondi',
                        'priority': 'high',
                        'raison': f"Taux de défaut de {r.get('taux_defaut', 0)}%",
                        'delai': 'Immédiat',
                        'impact': 'Réduction des défauts de 40-60%'
                    })
            elif niveau == 'Modéré':
                actions.append({
                    'supplier': supplier,
                    'action': 'Suivi mensuel des performances',
                    'priority': 'medium',
                    'raison': f"Score de risque: {r.get('score_risque', 0)}",
                    'delai': '1 mois',
                    'impact': 'Maintien de la qualité de service'
                })
    
    return actions


def calculate_case_specific_predictions(df: pd.DataFrame, data_type: DataTypeCase, fenetre: int = 3) -> List[Dict[str, Any]]:
    """
    Calculate predictions specific to the data type case.
    Uses existing ML model functions but filters output based on case.
    
    Case A: Only delay predictions
    Case B: Only defect predictions
    Case C: Both predictions
    """
    predictions = []
    
    for supplier in df['supplier'].unique():
        supplier_df = df[df['supplier'] == supplier]
        n_orders = len(supplier_df)
        
        pred = {
            'supplier': supplier,
            'nb_commandes_historique': n_orders,
            'confiance': 'haute' if n_orders >= 10 else 'moyenne' if n_orders >= 5 else 'basse'
        }
        
        if data_type == DataTypeCase.CASE_A:
            # ========================================
            # CASE A: DELAY PREDICTIONS ONLY
            # ========================================
            delays = supplier_df['delay'].values
            # Simple moving average prediction
            if len(delays) >= fenetre:
                pred['predicted_delay'] = round(float(delays[-fenetre:].mean()), 1)
            else:
                pred['predicted_delay'] = round(float(delays.mean()), 1)
            pred['predicted_defect'] = None  # Not applicable for Case A
            
        elif data_type == DataTypeCase.CASE_B:
            # ========================================
            # CASE B: DEFECT PREDICTIONS ONLY
            # ========================================
            defects = supplier_df['defects'].values
            # Simple moving average prediction
            if len(defects) >= fenetre:
                pred['predicted_defect'] = round(float(defects[-fenetre:].mean()) * 100, 2)
            else:
                pred['predicted_defect'] = round(float(defects.mean()) * 100, 2)
            pred['predicted_delay'] = None  # Not applicable for Case B
            
        elif data_type == DataTypeCase.CASE_C:
            # ========================================
            # CASE C: BOTH PREDICTIONS
            # ========================================
            delays = supplier_df['delay'].values
            defects = supplier_df['defects'].values
            
            # Delay prediction
            if len(delays) >= fenetre:
                pred['predicted_delay'] = round(float(delays[-fenetre:].mean()), 1)
            else:
                pred['predicted_delay'] = round(float(delays.mean()), 1)
            
            # Defect prediction
            if len(defects) >= fenetre:
                pred['predicted_defect'] = round(float(defects[-fenetre:].mean()) * 100, 2)
            else:
                pred['predicted_defect'] = round(float(defects.mean()) * 100, 2)
        
        predictions.append(pred)
    
    return predictions


@router.get("/{workspace_id}/analysis/dashboard")
async def get_workspace_dashboard(
    workspace_id: uuid.UUID,
    db: Session = Depends(get_db)
):
    """
    Get complete dashboard data for a workspace.
    Returns CASE-SPECIFIC KPIs, alerts, and predictions based on workspace data_type.
    
    Case A (delay_only): Delay KPIs, delay alerts, delay predictions
    Case B (defects_only): Defect KPIs, defect alerts, defect predictions  
    Case C (mixed): All KPIs, combined alerts, all predictions
    
    Uses EXISTING ML models without modification.
    """
    workspace = db.query(Workspace).filter(Workspace.id == workspace_id).first()
    if not workspace:
        raise HTTPException(status_code=404, detail="Workspace non trouvé")
    
    df = get_workspace_dataframe(workspace_id, db)
    if df is None or df.empty:
        raise HTTPException(status_code=400, detail="Aucune donnée disponible. Veuillez uploader un dataset.")
    
    # Validate required columns exist (delay and defects are always present after processing)
    required_cols = ['supplier', 'delay', 'defects']
    missing_cols = [col for col in required_cols if col not in df.columns]
    if missing_cols:
        raise HTTPException(
            status_code=400, 
            detail=f"Colonnes manquantes dans les données: {', '.join(missing_cols)}. Veuillez ré-uploader le dataset."
        )
    
    try:
        # Get model selection
        model_sel = db.query(ModelSelection).filter(
            ModelSelection.workspace_id == workspace_id
        ).first()
        
        fenetre = 3
        if model_sel and model_sel.parameters:
            fenetre = model_sel.parameters.get("fenetre", 3)
        
        # ========================================
        # CASE-SPECIFIC CALCULATIONS
        # Each case gets its own KPIs, risks, actions, and predictions
        # ========================================
        
        # Calculate case-specific KPIs
        kpis = calculate_case_specific_kpis(df, workspace.data_type)
        
        # Calculate case-specific supplier risks
        risques = calculate_case_specific_supplier_risks(df, workspace.data_type)
        
        # Calculate case-specific recommended actions
        actions = calculate_case_specific_actions(risques, workspace.data_type)
        
        # Calculate case-specific predictions
        predictions = calculate_case_specific_predictions(df, workspace.data_type, fenetre)
        
        # Calculate risk distribution
        distribution = {
            'faible': {'count': len([r for r in risques if r['niveau_risque'] == 'Faible']), 'label': 'Faible'},
            'modere': {'count': len([r for r in risques if r['niveau_risque'] == 'Modéré']), 'label': 'Modéré'},
            'eleve': {'count': len([r for r in risques if r['niveau_risque'] == 'Élevé']), 'label': 'Élevé'}
        }
        
        # Calculate custom KPIs
        custom_kpis = db.query(CustomKPI).filter(
            CustomKPI.workspace_id == workspace_id,
            CustomKPI.is_enabled == True
        ).all()
        
        custom_kpi_values = {}
        for kpi in custom_kpis:
            if kpi.target_field in df.columns:
                if kpi.formula_type == "average":
                    value = df[kpi.target_field].mean()
                elif kpi.formula_type == "sum":
                    value = df[kpi.target_field].sum()
                elif kpi.formula_type == "percentage":
                    value = (df[kpi.target_field] > 0).mean() * 100
                else:
                    value = df[kpi.target_field].mean()
                
                custom_kpi_values[kpi.name] = round(value, kpi.decimal_places)
        
        # Get case type description for frontend
        schema = get_schema_for_case(workspace.data_type)
        
        return {
            "workspace_id": str(workspace_id),
            "workspace_name": workspace.name,
            "data_type": workspace.data_type.value,
            "case_type": schema.get("case_type", "unknown"),
            "case_description": schema.get("description", ""),
            "kpis_globaux": kpis,
            "custom_kpis": custom_kpi_values,
            "suppliers": risques,
            "actions": actions,
            "predictions": predictions,
            "distribution": distribution,
            "selected_model": model_sel.selected_model if model_sel else "combined",
            "timestamp": datetime.now().isoformat()
        }
    except HTTPException:
        raise
    except Exception as e:
        # Log the actual error for debugging
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Erreur lors du calcul du dashboard: {str(e)}")


@router.get("/{workspace_id}/analysis/predictions")
async def get_workspace_predictions(
    workspace_id: uuid.UUID,
    supplier: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """
    Get predictions for a workspace, optionally filtered by supplier.
    Uses EXISTING prediction functions.
    """
    workspace = db.query(Workspace).filter(Workspace.id == workspace_id).first()
    if not workspace:
        raise HTTPException(status_code=404, detail="Workspace non trouvé")
    
    df = get_workspace_dataframe(workspace_id, db)
    if df is None or df.empty:
        raise HTTPException(status_code=400, detail="Aucune donnée disponible")
    
    # Get model selection parameters
    model_sel = db.query(ModelSelection).filter(
        ModelSelection.workspace_id == workspace_id
    ).first()
    
    fenetre = 3
    if model_sel and model_sel.parameters:
        fenetre = model_sel.parameters.get("fenetre", 3)
    
    # Use EXISTING prediction function
    predictions = calculer_predictions_avancees(df, fenetre=fenetre)
    
    # Filter by supplier if specified
    if supplier:
        predictions = [p for p in predictions if p["supplier"] == supplier]
    
    # If specific model selected, highlight that method's results
    selected_model = model_sel.selected_model if model_sel else "combined"
    
    return {
        "predictions": predictions,
        "selected_model": selected_model,
        "fenetre": fenetre,
        "filtered_by": supplier
    }


@router.get("/{workspace_id}/analysis/supplier/{supplier_name}")
async def get_supplier_analysis(
    workspace_id: uuid.UUID,
    supplier_name: str,
    db: Session = Depends(get_db)
):
    """
    Get detailed analysis for a specific supplier.
    Uses EXISTING analysis functions.
    """
    workspace = db.query(Workspace).filter(Workspace.id == workspace_id).first()
    if not workspace:
        raise HTTPException(status_code=404, detail="Workspace non trouvé")
    
    df = get_workspace_dataframe(workspace_id, db)
    if df is None or df.empty:
        raise HTTPException(status_code=400, detail="Aucune donnée disponible")
    
    # Use EXISTING comparison function
    comparison = comparer_methodes_prediction(df, supplier_name)
    if not comparison:
        raise HTTPException(status_code=404, detail=f"Fournisseur '{supplier_name}' non trouvé")
    
    # Get risk score for this supplier
    risques = calculer_risques_fournisseurs(df)
    supplier_risk = next((r for r in risques if r["supplier"] == supplier_name), None)
    
    # Get actions for this supplier
    actions = obtenir_actions_recommandees(risques)
    supplier_actions = [a for a in actions if a["supplier"] == supplier_name]
    
    return {
        "supplier": supplier_name,
        "risk_info": supplier_risk,
        "prediction_comparison": comparison,
        "recommended_actions": supplier_actions
    }


# ============================================
# MULTI-MODEL COMPARISON ENDPOINTS
# Run and compare predictions from multiple models
# ============================================

@router.get("/{workspace_id}/analysis/multi-model")
async def get_multi_model_predictions(
    workspace_id: uuid.UUID,
    models: str = Query("all", description="Comma-separated model IDs or 'all'"),
    supplier: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """
    Run predictions using multiple models and return side-by-side comparison.
    
    Models available: moving_average, linear_regression, exponential, combined
    
    Returns predictions from each selected model for comparison.
    """
    workspace = db.query(Workspace).filter(Workspace.id == workspace_id).first()
    if not workspace:
        raise HTTPException(status_code=404, detail="Workspace non trouvé")
    
    df = get_workspace_dataframe(workspace_id, db)
    if df is None or df.empty:
        raise HTTPException(status_code=400, detail="Aucune donnée disponible")
    
    # Get model parameters
    model_sel = db.query(ModelSelection).filter(
        ModelSelection.workspace_id == workspace_id
    ).first()
    
    fenetre = 3
    alpha = 0.3
    if model_sel and model_sel.parameters:
        fenetre = model_sel.parameters.get("fenetre", 3)
        alpha = model_sel.parameters.get("alpha", 0.3)
    
    # Determine which models to run
    available_models = ["moving_average", "linear_regression", "exponential", "combined"]
    if models == "all":
        selected_models = available_models
    else:
        selected_models = [m.strip() for m in models.split(",") if m.strip() in available_models]
    
    if not selected_models:
        raise HTTPException(status_code=400, detail="Aucun modèle valide sélectionné")
    
    # Filter by supplier if specified
    suppliers = [supplier] if supplier else df['supplier'].unique().tolist()
    
    # Calculate predictions for each model and supplier
    from backend.mon_analyse import (
        prediction_moyenne_mobile,
        prediction_regression_lineaire,
        prediction_lissage_exponentiel
    )
    
    results = []
    
    for sup in suppliers:
        supplier_df = df[df['supplier'] == sup]
        if len(supplier_df) < 2:
            continue
        
        delays = supplier_df['delay'].values
        defects = supplier_df['defects'].values
        n_orders = len(supplier_df)
        
        supplier_result = {
            "supplier": sup,
            "nb_commandes": n_orders,
            "predictions": {}
        }
        
        # Moving Average
        if "moving_average" in selected_models:
            ma_delay = prediction_moyenne_mobile(delays, fenetre)
            ma_defect = prediction_moyenne_mobile(defects * 100, fenetre)  # Convert to percentage
            supplier_result["predictions"]["moving_average"] = {
                "delay": round(ma_delay, 1) if ma_delay else None,
                "defect": round(ma_defect, 2) if ma_defect else None,
                "name": "Moyenne Glissante",
                "parameters": {"fenetre": fenetre}
            }
        
        # Linear Regression
        if "linear_regression" in selected_models:
            lr_delay = prediction_regression_lineaire(delays)
            lr_defect = prediction_regression_lineaire(defects * 100)
            supplier_result["predictions"]["linear_regression"] = {
                "delay": round(lr_delay, 1) if lr_delay else None,
                "defect": round(lr_defect, 2) if lr_defect else None,
                "name": "Régression Linéaire",
                "parameters": {}
            }
        
        # Exponential Smoothing
        if "exponential" in selected_models:
            exp_delay = prediction_lissage_exponentiel(delays, alpha)
            exp_defect = prediction_lissage_exponentiel(defects * 100, alpha)
            supplier_result["predictions"]["exponential"] = {
                "delay": round(exp_delay, 1) if exp_delay else None,
                "defect": round(exp_defect, 2) if exp_defect else None,
                "name": "Lissage Exponentiel",
                "parameters": {"alpha": alpha}
            }
        
        # Combined (average of all methods)
        if "combined" in selected_models:
            all_delays = []
            all_defects = []
            
            ma_d = prediction_moyenne_mobile(delays, fenetre)
            lr_d = prediction_regression_lineaire(delays)
            exp_d = prediction_lissage_exponentiel(delays, alpha)
            
            ma_def = prediction_moyenne_mobile(defects * 100, fenetre)
            lr_def = prediction_regression_lineaire(defects * 100)
            exp_def = prediction_lissage_exponentiel(defects * 100, alpha)
            
            for v in [ma_d, lr_d, exp_d]:
                if v is not None:
                    all_delays.append(v)
            for v in [ma_def, lr_def, exp_def]:
                if v is not None:
                    all_defects.append(v)
            
            combined_delay = sum(all_delays) / len(all_delays) if all_delays else None
            combined_defect = sum(all_defects) / len(all_defects) if all_defects else None
            
            supplier_result["predictions"]["combined"] = {
                "delay": round(combined_delay, 1) if combined_delay else None,
                "defect": round(combined_defect, 2) if combined_defect else None,
                "name": "Combiné (Moyenne)",
                "parameters": {"fenetre": fenetre, "alpha": alpha}
            }
        
        # Add case-specific filtering
        if workspace.data_type == DataTypeCase.CASE_A:
            # Delay only: null out defect predictions
            for model in supplier_result["predictions"].values():
                model["defect"] = None
        elif workspace.data_type == DataTypeCase.CASE_B:
            # Defects only: null out delay predictions
            for model in supplier_result["predictions"].values():
                model["delay"] = None
        
        results.append(supplier_result)
    
    return {
        "workspace_id": str(workspace_id),
        "case_type": workspace.data_type.value,
        "selected_models": selected_models,
        "parameters": {"fenetre": fenetre, "alpha": alpha},
        "results": results,
        "timestamp": datetime.now().isoformat()
    }


# ============================================
# EXPORT ENDPOINTS (PDF/Excel)
# ============================================

def _get_kpi_unit(kpi_name: str) -> str:
    """Helper function to get units for KPI indicators."""
    units = {
        'taux_retard': '%',
        'taux_defaut': '%',
        'taux_defauts': '%',
        'taux_conformite': '%',
        'retard_moyen': 'jours',
        'delai_moyen': 'jours',
        'nb_commandes': '',
        'commandes_parfaites': '',
        'score_global': '/100',
        'score_qualite': '/100',
        'score_ponctualite': '/100',
        'variabilite_defauts': '%',
        'variabilite_retards': '%',
    }
    return units.get(kpi_name.lower(), '')

@router.get("/{workspace_id}/export/excel")
async def export_to_excel(
    workspace_id: uuid.UUID,
    include_dashboard: bool = Query(True, description="Include dashboard KPIs"),
    include_predictions: bool = Query(True, description="Include predictions"),
    include_actions: bool = Query(True, description="Include recommended actions"),
    supplier: Optional[str] = Query(None, description="Filter by supplier"),
    db: Session = Depends(get_db)
):
    """
    Export workspace data, dashboard, and predictions to Excel format.
    Supports filtering by supplier.
    
    Returns a complete Excel workbook with multiple sheets:
    - Données: Raw normalized data
    - KPIs: Dashboard indicators
    - Risques Fournisseurs: Supplier risk scores
    - Prédictions: ML predictions per supplier
    - Actions Recommandées: Priority action items
    """
    # Pre-check: Verify openpyxl is installed before processing
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="Le module openpyxl n'est pas installé. Exécutez: pip install openpyxl"
        )
    
    workspace = db.query(Workspace).filter(Workspace.id == workspace_id).first()
    if not workspace:
        raise HTTPException(status_code=404, detail="Workspace non trouvé")
    
    df = get_workspace_dataframe(workspace_id, db)
    if df is None or df.empty:
        raise HTTPException(status_code=400, detail="Aucune donnée disponible pour ce workspace")
    
    try:
        # Filter by supplier if specified
        if supplier and supplier != 'all':
            df = df[df['supplier'] == supplier]
            if df.empty:
                raise HTTPException(status_code=404, detail=f"Fournisseur '{supplier}' non trouvé")
        
        # Create Excel file in memory
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Sheet 1: Workspace Info
            info_data = {
                'Propriété': ['Nom du Workspace', 'Type de Cas', 'Date d\'export', 'Fournisseur filtré', 'Nombre de lignes', 'Nombre de fournisseurs'],
                'Valeur': [
                    workspace.name,
                    workspace.data_type.value if workspace.data_type else 'N/A',
                    datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    supplier if supplier and supplier != 'all' else 'Tous',
                    len(df),
                    df['supplier'].nunique() if 'supplier' in df.columns else 0
                ]
            }
            pd.DataFrame(info_data).to_excel(writer, sheet_name='Informations', index=False)
            
            # Sheet 2: Raw Data
            df_export = df.copy()
            df_export.to_excel(writer, sheet_name='Données Normalisées', index=False)
            
            # Sheet 3: Dashboard KPIs
            if include_dashboard:
                kpis = calculate_case_specific_kpis(df, workspace.data_type)
                kpi_rows = [{"Indicateur": k, "Valeur": v, "Unité": _get_kpi_unit(k)} for k, v in kpis.items()]
                kpi_df = pd.DataFrame(kpi_rows)
                kpi_df.to_excel(writer, sheet_name='KPIs Dashboard', index=False)
            
            # Sheet 4: Supplier Risks
            risques = calculate_case_specific_supplier_risks(df, workspace.data_type)
            if supplier and supplier != 'all':
                risques = [r for r in risques if r.get('supplier') == supplier]
            if risques:
                risques_df = pd.DataFrame(risques)
                risques_df.to_excel(writer, sheet_name='Risques Fournisseurs', index=False)
            
            # Sheet 5: Predictions
            if include_predictions:
                predictions = calculate_case_specific_predictions(df, workspace.data_type)
                if supplier and supplier != 'all':
                    predictions = [p for p in predictions if p.get('supplier') == supplier]
                if predictions:
                    pred_df = pd.DataFrame(predictions)
                    pred_df.to_excel(writer, sheet_name='Prédictions', index=False)
            
            # Sheet 6: Recommended Actions
            if include_actions and risques:
                actions = calculate_case_specific_actions(risques, workspace.data_type)
                if supplier and supplier != 'all':
                    actions = [a for a in actions if a.get('supplier') == supplier]
                if actions:
                    actions_df = pd.DataFrame(actions)
                    actions_df.to_excel(writer, sheet_name='Actions Recommandées', index=False)
        
        output.seek(0)
        
        # Generate filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"workspace_{workspace.name}_{timestamp}"
        if supplier:
            filename += f"_{supplier}"
        filename += ".xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Erreur d'export: {str(e)}")


@router.get("/{workspace_id}/export/csv")
async def export_to_csv(
    workspace_id: uuid.UUID,
    data_type: str = Query("all", description="all, kpis, risks, predictions, actions"),
    supplier: Optional[str] = Query(None, description="Filter by supplier"),
    db: Session = Depends(get_db)
):
    """
    Export workspace data to CSV format.
    Choose what to export: raw data, KPIs, risks, predictions, or actions.
    """
    workspace = db.query(Workspace).filter(Workspace.id == workspace_id).first()
    if not workspace:
        raise HTTPException(status_code=404, detail="Workspace non trouvé")
    
    df = get_workspace_dataframe(workspace_id, db)
    if df is None or df.empty:
        raise HTTPException(status_code=400, detail="Aucune donnée disponible")
    
    try:
        # Filter by supplier if specified
        if supplier:
            df = df[df['supplier'] == supplier]
            if df.empty:
                raise HTTPException(status_code=404, detail=f"Fournisseur '{supplier}' non trouvé")
        
        output = io.StringIO()
        
        if data_type == "all" or data_type == "data":
            df.to_csv(output, index=False)
        elif data_type == "kpis":
            kpis = calculate_case_specific_kpis(df, workspace.data_type)
            kpi_df = pd.DataFrame([{"KPI": k, "Valeur": v} for k, v in kpis.items()])
            kpi_df.to_csv(output, index=False)
        elif data_type == "risks":
            risques = calculate_case_specific_supplier_risks(df, workspace.data_type)
            if supplier:
                risques = [r for r in risques if r['supplier'] == supplier]
            pd.DataFrame(risques).to_csv(output, index=False)
        elif data_type == "predictions":
            predictions = calculate_case_specific_predictions(df, workspace.data_type)
            if supplier:
                predictions = [p for p in predictions if p['supplier'] == supplier]
            pd.DataFrame(predictions).to_csv(output, index=False)
        elif data_type == "actions":
            risques = calculate_case_specific_supplier_risks(df, workspace.data_type)
            actions = calculate_case_specific_actions(risques, workspace.data_type)
            if supplier:
                actions = [a for a in actions if a['supplier'] == supplier]
            pd.DataFrame(actions).to_csv(output, index=False)
        else:
            raise HTTPException(status_code=400, detail="Type d'export invalide")
        
        output.seek(0)
        
        # Generate filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"workspace_{workspace.name}_{data_type}_{timestamp}"
        if supplier:
            filename += f"_{supplier}"
        filename += ".csv"
        
        return StreamingResponse(
            io.BytesIO(output.getvalue().encode('utf-8')),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur d'export: {str(e)}")


@router.get("/{workspace_id}/export/report")
async def export_report_summary(
    workspace_id: uuid.UUID,
    supplier: Optional[str] = Query(None, description="Filter by supplier"),
    db: Session = Depends(get_db)
):
    """
    Export a summary report in JSON format.
    Can be used to generate PDF reports on the frontend.
    """
    workspace = db.query(Workspace).filter(Workspace.id == workspace_id).first()
    if not workspace:
        raise HTTPException(status_code=404, detail="Workspace non trouvé")
    
    df = get_workspace_dataframe(workspace_id, db)
    if df is None or df.empty:
        raise HTTPException(status_code=400, detail="Aucune donnée disponible")
    
    try:
        # Filter by supplier if specified
        if supplier:
            df_filtered = df[df['supplier'] == supplier]
            if df_filtered.empty:
                raise HTTPException(status_code=404, detail=f"Fournisseur '{supplier}' non trouvé")
        else:
            df_filtered = df
        
        # Calculate all data
        kpis = calculate_case_specific_kpis(df_filtered, workspace.data_type)
        risques = calculate_case_specific_supplier_risks(df_filtered, workspace.data_type)
        actions = calculate_case_specific_actions(risques, workspace.data_type)
        predictions = calculate_case_specific_predictions(df_filtered, workspace.data_type)
        
        if supplier:
            risques = [r for r in risques if r['supplier'] == supplier]
            actions = [a for a in actions if a['supplier'] == supplier]
            predictions = [p for p in predictions if p['supplier'] == supplier]
        
        # Get schema info
        schema = get_schema_for_case(workspace.data_type)
        
        report = {
            "report_info": {
                "workspace_name": workspace.name,
                "workspace_id": str(workspace_id),
                "generated_at": datetime.now().isoformat(),
                "filtered_by_supplier": supplier,
                "case_type": schema.get("case_type"),
                "case_description": schema.get("description")
            },
            "data_summary": {
                "total_rows": len(df_filtered),
                "total_suppliers": df_filtered['supplier'].nunique(),
                "suppliers": df_filtered['supplier'].unique().tolist(),
                "date_range": {
                    "start": df_filtered['date_promised'].min().strftime("%Y-%m-%d") if 'date_promised' in df_filtered.columns else None,
                    "end": df_filtered['date_promised'].max().strftime("%Y-%m-%d") if 'date_promised' in df_filtered.columns else None
                }
            },
            "kpis": kpis,
            "risk_distribution": {
                "faible": len([r for r in risques if r.get('niveau_risque') == 'Faible']),
                "modere": len([r for r in risques if r.get('niveau_risque') == 'Modéré']),
                "eleve": len([r for r in risques if r.get('niveau_risque') == 'Élevé'])
            },
            "supplier_risks": risques,
            "predictions": predictions,
            "recommended_actions": {
                "high_priority": [a for a in actions if a.get('priority') == 'high'],
                "medium_priority": [a for a in actions if a.get('priority') == 'medium'],
                "low_priority": [a for a in actions if a.get('priority') == 'low']
            }
        }
        
        return report
        
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Erreur de génération du rapport: {str(e)}")


# ============================================
# SUPPLIER MANAGEMENT ENDPOINTS
# Manage suppliers within a workspace
# ============================================

class SupplierCreate(BaseModel):
    """Schema for creating/adding a supplier to a workspace"""
    name: str = Field(..., min_length=1, max_length=200)
    description: Optional[str] = None
    category: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None


class SupplierUpdate(BaseModel):
    """Schema for updating supplier metadata"""
    name: Optional[str] = Field(None, min_length=1, max_length=200)
    description: Optional[str] = None
    category: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None


class OrderCreate(BaseModel):
    """Schema for manually creating an order record"""
    supplier_name: str = Field(..., min_length=1)
    date_promised: Optional[str] = Field(None, description="Date format: YYYY-MM-DD (required for Case A and C)")
    date_delivered: Optional[str] = Field(None, description="Date format: YYYY-MM-DD")
    defects: Optional[float] = Field(None, ge=0.0, le=1.0, description="Defect rate 0-1 (required for Case B)")
    order_reference: Optional[str] = None
    quantity: Optional[int] = None
    amount: Optional[float] = None
    notes: Optional[str] = None


class BulkOrderCreate(BaseModel):
    """Schema for adding multiple orders at once"""
    orders: List[OrderCreate]


@router.get("/{workspace_id}/suppliers", response_model=Dict[str, Any])
async def list_workspace_suppliers(
    workspace_id: uuid.UUID,
    db: Session = Depends(get_db)
):
    """
    List all suppliers in a workspace with their statistics.
    Data comes from the workspace dataset.
    """
    workspace = db.query(Workspace).filter(Workspace.id == workspace_id).first()
    if not workspace:
        raise HTTPException(status_code=404, detail="Workspace non trouvé")
    
    df = get_workspace_dataframe(workspace_id, db)
    
    if df is None or df.empty:
        return {
            "suppliers": [],
            "total_count": 0,
            "message": "Aucun fournisseur. Importez des données pour commencer."
        }
    
    # Get supplier statistics
    suppliers_data = []
    for supplier_name in df['supplier'].unique():
        supplier_df = df[df['supplier'] == supplier_name]
        
        stats = {
            "name": supplier_name,
            "order_count": len(supplier_df),
            "first_order": supplier_df['date_promised'].min().strftime("%Y-%m-%d") if 'date_promised' in supplier_df.columns else None,
            "last_order": supplier_df['date_promised'].max().strftime("%Y-%m-%d") if 'date_promised' in supplier_df.columns else None,
        }
        
        # Add case-specific stats
        if workspace.data_type in [DataTypeCase.CASE_A, DataTypeCase.CASE_C]:
            if 'delay' in supplier_df.columns:
                stats["avg_delay"] = round(supplier_df['delay'].mean(), 2)
                stats["max_delay"] = int(supplier_df['delay'].max())
                stats["on_time_rate"] = round((supplier_df['delay'] == 0).sum() / len(supplier_df) * 100, 1)
        
        if workspace.data_type in [DataTypeCase.CASE_B, DataTypeCase.CASE_C]:
            if 'defects' in supplier_df.columns:
                stats["avg_defects"] = round(supplier_df['defects'].mean() * 100, 2)
                stats["max_defects"] = round(supplier_df['defects'].max() * 100, 2)
        
        suppliers_data.append(stats)
    
    # Sort by order count descending
    suppliers_data.sort(key=lambda x: x['order_count'], reverse=True)
    
    return {
        "suppliers": suppliers_data,
        "total_count": len(suppliers_data),
        "workspace_data_type": workspace.data_type.value
    }


@router.post("/{workspace_id}/suppliers", response_model=Dict[str, Any])
async def add_supplier(
    workspace_id: uuid.UUID,
    supplier: SupplierCreate,
    db: Session = Depends(get_db)
):
    """
    Add a new supplier to a workspace.
    Creates an empty record - orders need to be added separately.
    """
    workspace = db.query(Workspace).filter(Workspace.id == workspace_id).first()
    if not workspace:
        raise HTTPException(status_code=404, detail="Workspace non trouvé")
    
    # Get current dataset
    dataset = db.query(WorkspaceDataset).filter(
        WorkspaceDataset.workspace_id == workspace_id,
        WorkspaceDataset.is_active == True
    ).first()
    
    if dataset:
        # Check if supplier already exists
        if supplier.name in dataset.suppliers:
            raise HTTPException(status_code=400, detail=f"Le fournisseur '{supplier.name}' existe déjà")
        
        # Add to suppliers list
        new_suppliers = dataset.suppliers + [supplier.name]
        dataset.suppliers = new_suppliers
        db.commit()
    else:
        # Create a new empty dataset with just the supplier
        new_dataset = WorkspaceDataset(
            workspace_id=workspace_id,
            filename="manual_entry.csv",
            row_count=0,
            column_count=4,
            suppliers=[supplier.name],
            data_json=[],
            is_active=True
        )
        db.add(new_dataset)
        db.commit()
        db.refresh(new_dataset)
    
    return {
        "success": True,
        "message": f"Fournisseur '{supplier.name}' ajouté avec succès",
        "supplier": {
            "name": supplier.name,
            "description": supplier.description,
            "category": supplier.category
        }
    }


@router.delete("/{workspace_id}/suppliers/{supplier_name}", response_model=Dict[str, Any])
async def remove_supplier(
    workspace_id: uuid.UUID,
    supplier_name: str,
    db: Session = Depends(get_db)
):
    """
    Remove a supplier and all their orders from the workspace.
    This is a destructive operation!
    """
    workspace = db.query(Workspace).filter(Workspace.id == workspace_id).first()
    if not workspace:
        raise HTTPException(status_code=404, detail="Workspace non trouvé")
    
    dataset = db.query(WorkspaceDataset).filter(
        WorkspaceDataset.workspace_id == workspace_id,
        WorkspaceDataset.is_active == True
    ).first()
    
    if not dataset:
        raise HTTPException(status_code=400, detail="Aucune donnée dans ce workspace")
    
    # Check if supplier exists
    if supplier_name not in dataset.suppliers:
        raise HTTPException(status_code=404, detail=f"Fournisseur '{supplier_name}' non trouvé")
    
    # Remove supplier from list
    new_suppliers = [s for s in dataset.suppliers if s != supplier_name]
    
    # Remove orders from data
    data = dataset.data_json or []
    new_data = [row for row in data if row.get('supplier') != supplier_name]
    
    # Update dataset
    dataset.suppliers = new_suppliers
    dataset.data_json = new_data
    dataset.row_count = len(new_data)
    
    db.commit()
    
    return {
        "success": True,
        "message": f"Fournisseur '{supplier_name}' supprimé avec succès",
        "removed_orders": len(data) - len(new_data),
        "remaining_suppliers": len(new_suppliers)
    }


# ============================================
# ORDER MANAGEMENT ENDPOINTS
# Add/manage orders within a workspace
# ============================================

@router.get("/{workspace_id}/suppliers/{supplier_name}/orders", response_model=Dict[str, Any])
async def get_supplier_orders(
    workspace_id: uuid.UUID,
    supplier_name: str,
    limit: int = Query(100, ge=1, le=1000),
    offset: int = Query(0, ge=0),
    db: Session = Depends(get_db)
):
    """
    Get orders for a specific supplier in the workspace.
    """
    workspace = db.query(Workspace).filter(Workspace.id == workspace_id).first()
    if not workspace:
        raise HTTPException(status_code=404, detail="Workspace non trouvé")
    
    df = get_workspace_dataframe(workspace_id, db)
    
    if df is None or df.empty:
        raise HTTPException(status_code=400, detail="Aucune donnée disponible")
    
    # Filter by supplier
    supplier_df = df[df['supplier'] == supplier_name]
    
    if supplier_df.empty:
        raise HTTPException(status_code=404, detail=f"Fournisseur '{supplier_name}' non trouvé")
    
    # Paginate
    total = len(supplier_df)
    supplier_df = supplier_df.iloc[offset:offset + limit]
    
    # Convert to list of dicts
    orders = []
    for _, row in supplier_df.iterrows():
        order = {
            "supplier": row['supplier'],
            "date_promised": row['date_promised'].strftime("%Y-%m-%d") if pd.notna(row.get('date_promised')) else None,
            "date_delivered": row['date_delivered'].strftime("%Y-%m-%d") if pd.notna(row.get('date_delivered')) else None,
        }
        
        if 'delay' in row:
            order['delay'] = int(row['delay']) if pd.notna(row['delay']) else 0
        if 'defects' in row:
            order['defects'] = round(float(row['defects']), 4) if pd.notna(row['defects']) else 0.0
        if 'order_reference' in row:
            order['order_reference'] = row['order_reference']
        if 'quantity' in row:
            order['quantity'] = int(row['quantity']) if pd.notna(row['quantity']) else None
        if 'amount' in row:
            order['amount'] = float(row['amount']) if pd.notna(row['amount']) else None
        
        orders.append(order)
    
    return {
        "supplier": supplier_name,
        "orders": orders,
        "total": total,
        "limit": limit,
        "offset": offset,
        "has_more": offset + limit < total
    }


@router.post("/{workspace_id}/orders", response_model=Dict[str, Any])
async def add_manual_order(
    workspace_id: uuid.UUID,
    order: OrderCreate,
    db: Session = Depends(get_db)
):
    """
    Add a single order manually to the workspace.
    The order will be validated based on workspace data_type:
    - Case A (delays): requires date_promised, date_delivered optional
    - Case B (late_days): requires defects
    - Case C (mixed): requires date_promised and defects
    
    After successful insertion:
    - Dataset is updated with the new order
    - Frontend should refetch dashboard to reflect updated KPIs, predictions, and charts
    """
    workspace = db.query(Workspace).filter(Workspace.id == workspace_id).first()
    if not workspace:
        raise HTTPException(status_code=404, detail="Workspace non trouvé")
    
    data_type = workspace.data_type or "delays"
    
    # Case-specific validation
    date_promised = None
    date_delivered = None
    delay = 0
    
    if data_type in ["delays", "mixed"]:
        # Case A and C require dates
        if not order.date_promised:
            raise HTTPException(status_code=400, detail="Date promise requise pour ce type de workspace")
        try:
            date_promised = datetime.strptime(order.date_promised, "%Y-%m-%d")
            date_delivered = datetime.strptime(order.date_delivered, "%Y-%m-%d") if order.date_delivered else date_promised
        except ValueError:
            raise HTTPException(status_code=400, detail="Format de date invalide. Utilisez YYYY-MM-DD")
        # Calculate delay
        delay = max((date_delivered - date_promised).days, 0)
    
    if data_type == "late_days":
        # Case B requires defects
        if order.defects is None:
            raise HTTPException(status_code=400, detail="Taux de défauts requis pour ce type de workspace")
    
    try:
        # Get current dataset
        dataset = db.query(WorkspaceDataset).filter(
            WorkspaceDataset.workspace_id == workspace_id,
            WorkspaceDataset.is_active == True
        ).first()
        
        # Create order record based on case type
        new_order = {
            "supplier": order.supplier_name,
        }
        
        # Add date fields for Case A and C
        if data_type in ["delays", "mixed"]:
            new_order["date_promised"] = order.date_promised
            new_order["date_delivered"] = order.date_delivered or order.date_promised
            new_order["delay"] = delay
        
        # Add defects field for Case B and C
        if data_type in ["late_days", "mixed"]:
            new_order["defects"] = order.defects if order.defects is not None else 0.0
        
        # Add optional fields
        if order.order_reference:
            new_order["order_reference"] = order.order_reference
        if order.quantity is not None:
            new_order["quantity"] = order.quantity
        if order.amount is not None:
            new_order["amount"] = order.amount
        if order.notes:
            new_order["notes"] = order.notes
        
        new_dataset = None
        if dataset:
            # Merge with existing data
            # IMPORTANT: Create a NEW list to ensure SQLAlchemy detects the change
            # SQLAlchemy doesn't detect in-place mutations of JSON columns
            data = list(dataset.data_json or [])
            data.append(new_order)
            
            # Update suppliers list if new supplier (also create new list)
            suppliers = list(dataset.suppliers or [])
            if order.supplier_name not in suppliers:
                suppliers.append(order.supplier_name)
            
            # Update date range (only for Case A and C which have dates)
            date_start = dataset.date_start
            date_end = dataset.date_end
            
            if order.date_promised and data_type in ["delays", "mixed"]:
                new_date = datetime.strptime(order.date_promised, "%Y-%m-%d")
                # Ensure date comparison works by making dates naive (remove timezone info if present)
                if date_start is not None and hasattr(date_start, 'tzinfo') and date_start.tzinfo is not None:
                    date_start = date_start.replace(tzinfo=None)
                if date_end is not None and hasattr(date_end, 'tzinfo') and date_end.tzinfo is not None:
                    date_end = date_end.replace(tzinfo=None)
                
                if date_start is None or new_date < date_start:
                    date_start = new_date
                if date_end is None or new_date > date_end:
                    date_end = new_date
            
            # Update dataset - assign new lists to trigger SQLAlchemy change detection
            dataset.data_json = data
            dataset.row_count = len(data)
            dataset.suppliers = suppliers
            dataset.date_start = date_start
            dataset.date_end = date_end
            
            # Explicitly mark JSON columns as modified to ensure SQLAlchemy persists them
            flag_modified(dataset, 'data_json')
            flag_modified(dataset, 'suppliers')
            
            db.commit()
        else:
            # Create new dataset
            new_dataset = WorkspaceDataset(
                workspace_id=workspace_id,
                filename="manual_entry.csv",
                row_count=1,
                column_count=len(new_order),
                suppliers=[order.supplier_name],
                date_start=date_promised,  # Will be None for Case B
                date_end=date_promised,    # Will be None for Case B
                data_json=[new_order],
                is_active=True
            )
            db.add(new_dataset)
            db.commit()
        
        # Refresh dataset reference after commit to get updated row count
        db.refresh(dataset if dataset else new_dataset)
        updated_dataset = dataset if dataset else new_dataset
        
        return {
            "success": True,
            "message": "Commande ajoutée avec succès",
            "order": new_order,
            "dataset_updated": True,
            "new_row_count": updated_dataset.row_count,
            "refresh_required": True,  # Signal frontend to refresh dashboard
            "supplier_updated": order.supplier_name
        }
    except HTTPException:
        raise
    except Exception as e:
        # Rollback any partial changes on error
        db.rollback()
        import traceback
        traceback.print_exc()
        raise HTTPException(
            status_code=500, 
            detail=f"Erreur lors de l'ajout de la commande: {str(e)}. Les modifications ont été annulées."
        )


@router.post("/{workspace_id}/orders/bulk", response_model=Dict[str, Any])
async def add_bulk_orders(
    workspace_id: uuid.UUID,
    bulk_data: BulkOrderCreate,
    db: Session = Depends(get_db)
):
    """
    Add multiple orders at once to the workspace.
    All orders are validated and merged with existing data.
    """
    workspace = db.query(Workspace).filter(Workspace.id == workspace_id).first()
    if not workspace:
        raise HTTPException(status_code=404, detail="Workspace non trouvé")
    
    if not bulk_data.orders:
        raise HTTPException(status_code=400, detail="Aucune commande fournie")
    
    # Get current dataset
    dataset = db.query(WorkspaceDataset).filter(
        WorkspaceDataset.workspace_id == workspace_id,
        WorkspaceDataset.is_active == True
    ).first()
    
    # Process orders
    new_orders = []
    errors = []
    new_suppliers = set()
    
    for i, order in enumerate(bulk_data.orders):
        try:
            # Validate dates
            date_promised = datetime.strptime(order.date_promised, "%Y-%m-%d")
            date_delivered = datetime.strptime(order.date_delivered, "%Y-%m-%d") if order.date_delivered else date_promised
            delay = max((date_delivered - date_promised).days, 0)
            
            new_order = {
                "supplier": order.supplier_name,
                "date_promised": order.date_promised,
                "date_delivered": order.date_delivered or order.date_promised,
                "delay": delay,
                "defects": order.defects or 0.0
            }
            
            if order.order_reference:
                new_order["order_reference"] = order.order_reference
            if order.quantity is not None:
                new_order["quantity"] = order.quantity
            if order.amount is not None:
                new_order["amount"] = order.amount
            
            new_orders.append(new_order)
            new_suppliers.add(order.supplier_name)
            
        except ValueError as e:
            errors.append(f"Ligne {i + 1}: {str(e)}")
    
    if not new_orders:
        raise HTTPException(status_code=400, detail={"message": "Aucune commande valide", "errors": errors})
    
    # Merge with existing data
    if dataset:
        # Create NEW list to ensure SQLAlchemy detects the change
        data = list(dataset.data_json or [])
        data.extend(new_orders)
        
        suppliers = set(dataset.suppliers or [])
        suppliers.update(new_suppliers)
        
        # Recalculate date range
        all_dates = [datetime.strptime(o['date_promised'], "%Y-%m-%d") for o in data]
        date_start = min(all_dates) if all_dates else None
        date_end = max(all_dates) if all_dates else None
        
        dataset.data_json = data
        dataset.row_count = len(data)
        dataset.suppliers = list(suppliers)
        dataset.date_start = date_start
        dataset.date_end = date_end
        
        # Explicitly mark JSON columns as modified
        flag_modified(dataset, 'data_json')
        flag_modified(dataset, 'suppliers')
        
        db.commit()
    else:
        # Calculate date range
        all_dates = [datetime.strptime(o['date_promised'], "%Y-%m-%d") for o in new_orders]
        
        new_dataset = WorkspaceDataset(
            workspace_id=workspace_id,
            filename="manual_entry.csv",
            row_count=len(new_orders),
            column_count=5,
            suppliers=list(new_suppliers),
            date_start=min(all_dates) if all_dates else None,
            date_end=max(all_dates) if all_dates else None,
            data_json=new_orders,
            is_active=True
        )
        db.add(new_dataset)
        db.commit()
    
    # Refresh dataset reference after commit
    db.refresh(dataset if dataset else new_dataset)
    updated_dataset = dataset if dataset else new_dataset
    
    return {
        "success": True,
        "message": f"{len(new_orders)} commandes ajoutées avec succès",
        "added_count": len(new_orders),
        "error_count": len(errors),
        "errors": errors if errors else None,
        "new_suppliers": list(new_suppliers),
        "dataset_updated": True,
        "new_row_count": updated_dataset.row_count,
        "refresh_required": True  # Signal frontend to refresh dashboard
    }


@router.post("/{workspace_id}/suppliers/{supplier_name}/upload", response_model=Dict[str, Any])
async def upload_supplier_csv(
    workspace_id: uuid.UUID,
    supplier_name: str,
    file: UploadFile = File(...),
    merge_mode: str = Query("append", description="'append' or 'replace'"),
    db: Session = Depends(get_db)
):
    """
    Upload CSV data for a specific supplier.
    Data is validated and merged with existing workspace data.
    
    merge_mode:
    - 'append': Add new orders to existing data
    - 'replace': Replace all orders for this supplier
    """
    workspace = db.query(Workspace).filter(Workspace.id == workspace_id).first()
    if not workspace:
        raise HTTPException(status_code=404, detail="Workspace non trouvé")
    
    if not file.filename.endswith('.csv'):
        raise HTTPException(status_code=400, detail="Format invalide. Fichier CSV requis.")
    
    try:
        # Read CSV
        content = await file.read()
        df = pd.read_csv(io.BytesIO(content))
        
        if df.empty:
            raise HTTPException(status_code=400, detail="Le fichier CSV est vide.")
        
        # Override supplier column with the specified supplier name
        df['supplier'] = supplier_name
        
        # Validate against workspace data type
        errors = validate_csv_for_case(df, workspace.data_type)
        if errors:
            raise HTTPException(
                status_code=400,
                detail={"message": "Erreurs de validation", "errors": errors}
            )
        
        # Process data
        processed_df = process_csv_for_case(df, workspace.data_type)
        
        # Get current dataset
        dataset = db.query(WorkspaceDataset).filter(
            WorkspaceDataset.workspace_id == workspace_id,
            WorkspaceDataset.is_active == True
        ).first()
        
        # Convert processed data to list of dicts
        new_orders = processed_df.to_dict(orient='records')
        for order in new_orders:
            # Convert dates to strings
            for col in ['date_promised', 'date_delivered', 'order_date']:
                if col in order and pd.notna(order[col]):
                    if isinstance(order[col], (datetime, pd.Timestamp)):
                        order[col] = order[col].strftime("%Y-%m-%d")
        
        if dataset:
            # Create NEW list to ensure SQLAlchemy detects the change
            existing_data = list(dataset.data_json or [])
            
            if merge_mode == "replace":
                # Remove existing orders for this supplier
                existing_data = [o for o in existing_data if o.get('supplier') != supplier_name]
            
            # Add new orders
            existing_data.extend(new_orders)
            
            # Update suppliers list
            suppliers = set(dataset.suppliers or [])
            suppliers.add(supplier_name)
            
            # Recalculate date range
            date_col = "date_promised" if any('date_promised' in o for o in existing_data) else "order_date"
            all_dates = []
            for o in existing_data:
                if date_col in o and o[date_col]:
                    try:
                        all_dates.append(datetime.strptime(o[date_col], "%Y-%m-%d"))
                    except:
                        pass
            
            dataset.data_json = existing_data
            dataset.row_count = len(existing_data)
            dataset.suppliers = list(suppliers)
            dataset.date_start = min(all_dates) if all_dates else None
            dataset.date_end = max(all_dates) if all_dates else None
            
            # Explicitly mark JSON columns as modified
            flag_modified(dataset, 'data_json')
            flag_modified(dataset, 'suppliers')
            
            db.commit()
        else:
            # Create new dataset
            date_col = "date_promised" if "date_promised" in processed_df.columns else "order_date"
            date_start = processed_df[date_col].min() if date_col in processed_df.columns else None
            date_end = processed_df[date_col].max() if date_col in processed_df.columns else None
            
            new_dataset = WorkspaceDataset(
                workspace_id=workspace_id,
                filename=f"{supplier_name}_{file.filename}",
                row_count=len(new_orders),
                column_count=len(processed_df.columns),
                suppliers=[supplier_name],
                date_start=date_start.to_pydatetime() if pd.notna(date_start) else None,
                date_end=date_end.to_pydatetime() if pd.notna(date_end) else None,
                data_json=new_orders,
                is_active=True
            )
            db.add(new_dataset)
            db.commit()
        
        # Refresh to get updated data
        db.refresh(dataset if dataset else new_dataset)
        updated_dataset = dataset if dataset else new_dataset
        
        return {
            "success": True,
            "message": f"{len(new_orders)} commandes importées pour {supplier_name}",
            "supplier": supplier_name,
            "orders_added": len(new_orders),
            "merge_mode": merge_mode,
            "dataset_updated": True,
            "new_row_count": updated_dataset.row_count,
            "refresh_required": True  # Signal frontend to refresh dashboard
        }
        
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Erreur de traitement: {str(e)}")


@router.post("/{workspace_id}/suppliers/{supplier_name}/upload/smart", response_model=Dict[str, Any])
async def smart_upload_supplier_csv(
    workspace_id: uuid.UUID,
    supplier_name: str,
    file: UploadFile = File(...),
    merge_mode: str = Query("append", description="'append' or 'replace'"),
    db: Session = Depends(get_db)
):
    """
    Smart upload CSV data for a specific supplier using LLM-based column mapping.
    Automatically detects and maps columns to the expected schema.
    """
    workspace = db.query(Workspace).filter(Workspace.id == workspace_id).first()
    if not workspace:
        raise HTTPException(status_code=404, detail="Workspace non trouvé")
    
    if not file.filename.endswith('.csv'):
        raise HTTPException(status_code=400, detail="Format invalide. Fichier CSV requis.")
    
    try:
        # Read CSV
        content = await file.read()
        df = pd.read_csv(io.BytesIO(content))
        
        if df.empty:
            raise HTTPException(status_code=400, detail="Le fichier CSV est vide.")
        
        # Analyze CSV with LLM-style detection
        analysis = analyze_csv_for_mapping(df)
        
        # Check confidence
        all_high_confidence = all(
            m["confidence"] > 0.7 
            for m in analysis["mappings"] 
            if m["target_role"] != "ignore"
        )
        
        if not all_high_confidence:
            # Return analysis for manual review
            import base64
            return {
                "success": False,
                "needs_review": True,
                "message": "Certaines colonnes nécessitent une vérification manuelle",
                "analysis": analysis,
                "csv_content": base64.b64encode(content).decode('utf-8')
            }
        
        # Apply mappings
        target_case = {
            DataTypeCase.CASE_A: "delay_only",
            DataTypeCase.CASE_B: "defects_only",
            DataTypeCase.CASE_C: "mixed"
        }.get(workspace.data_type, "mixed")
        
        result = process_csv_with_llm_mapping(
            df, 
            user_mappings=analysis["mappings"],
            target_case=target_case
        )
        
        if not result.success:
            return {
                "success": False,
                "message": "Échec de la normalisation",
                "errors": [w.message for w in result.warnings if w.severity == "error"]
            }
        
        # Override supplier column
        processed_df = result.dataframe
        processed_df['supplier'] = supplier_name
        
        # Get current dataset
        dataset = db.query(WorkspaceDataset).filter(
            WorkspaceDataset.workspace_id == workspace_id,
            WorkspaceDataset.is_active == True
        ).first()
        
        # Convert to list of dicts
        new_orders = processed_df.to_dict(orient='records')
        for order in new_orders:
            for col in ['date_promised', 'date_delivered', 'order_date']:
                if col in order and pd.notna(order[col]):
                    if isinstance(order[col], (datetime, pd.Timestamp)):
                        order[col] = order[col].strftime("%Y-%m-%d")
        
        if dataset:
            # Create NEW list to ensure SQLAlchemy detects the change
            existing_data = list(dataset.data_json or [])
            
            if merge_mode == "replace":
                existing_data = [o for o in existing_data if o.get('supplier') != supplier_name]
            
            existing_data.extend(new_orders)
            
            suppliers = set(dataset.suppliers or [])
            suppliers.add(supplier_name)
            
            # Recalculate date range
            date_col = "date_promised"
            all_dates = []
            for o in existing_data:
                if date_col in o and o[date_col]:
                    try:
                        all_dates.append(datetime.strptime(o[date_col], "%Y-%m-%d"))
                    except:
                        pass
            
            dataset.data_json = existing_data
            dataset.row_count = len(existing_data)
            dataset.suppliers = list(suppliers)
            dataset.date_start = min(all_dates) if all_dates else None
            dataset.date_end = max(all_dates) if all_dates else None
            
            # Explicitly mark JSON columns as modified
            flag_modified(dataset, 'data_json')
            flag_modified(dataset, 'suppliers')
            
            db.commit()
        else:
            date_col = "date_promised"
            all_dates = [datetime.strptime(o[date_col], "%Y-%m-%d") for o in new_orders if date_col in o and o[date_col]]
            
            new_dataset = WorkspaceDataset(
                workspace_id=workspace_id,
                filename=f"{supplier_name}_{file.filename}",
                row_count=len(new_orders),
                column_count=len(processed_df.columns),
                suppliers=[supplier_name],
                date_start=min(all_dates) if all_dates else None,
                date_end=max(all_dates) if all_dates else None,
                data_json=new_orders,
                is_active=True
            )
            db.add(new_dataset)
            db.commit()
        
        # Refresh to get updated data
        db.refresh(dataset if dataset else new_dataset)
        updated_dataset = dataset if dataset else new_dataset
        
        return {
            "success": True,
            "message": f"{len(new_orders)} commandes importées intelligemment pour {supplier_name}",
            "supplier": supplier_name,
            "orders_added": len(new_orders),
            "merge_mode": merge_mode,
            "detected_mappings": analysis["mappings"],
            "dataset_updated": True,
            "new_row_count": updated_dataset.row_count,
            "refresh_required": True  # Signal frontend to refresh dashboard
        }
        
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Erreur de traitement: {str(e)}")

